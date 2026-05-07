import re
import PyPDF2
import pandas as pd
from tkinter import filedialog
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import numbers, Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter

# ===================== KONFIGURASI FORMAT =====================
TEXT_HEADERS = {
    "NOMOR_FAKTUR",
    "REFERENSI",
    "NPWP_PENJUAL",
    "NPWP_PEMBELI",
    "KODE_BARANG",
    "NIK",
    "NOMOR PASPOR",
    "IDENTITAS LAIN",
    "EMAIL"
}

DATE_HEADERS = {
    "TANGGAL"
}

NUMBER_HEADERS = {
    "HARGA",
    "QTY",
    "POTONGAN",
    "PPNBM",
    "TOTAL_HARGA_BARANG",
    "TOTAL_HARGA",
    "TOTAL_POTONGAN",
    "DPP",
    "PPN",
    "TOTAL_PPNBM",
    "NO"  # Tambahan untuk kolom nomor urut
}

# Kolom yang akan di-merge (data header faktur) + NO
MERGE_COLUMNS = {
    "NO",  # Kolom nomor urut juga di-merge
    "NOMOR_FAKTUR",
    "REFERENSI",
    "KOTA",
    "TANGGAL",
    "PENANDATANGAN",
    "NAMA_PENJUAL",
    "ALAMAT_PENJUAL",
    "NPWP_PENJUAL",
    "NPWP_PEMBELI",
    "NAMA_PEMBELI",
    "ALAMAT_PEMBELI",
    "NIK",
    "NOMOR PASPOR",
    "IDENTITAS LAIN",
    "EMAIL",
    "TOTAL_HARGA",
    "TOTAL_POTONGAN",
    "DPP",
    "PPN",
    "TOTAL_PPNBM"
}

BULAN_MAP = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4,
    "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
    "september": 9, "oktober": 10, "november": 11, "desember": 12
}

# ===================== HELPER FUNCTIONS =====================
def parse_tanggal_teks(tanggal_str):
    """Convert '20 Desember 2025' to datetime object"""
    if not tanggal_str or tanggal_str == "":
        return None
    
    try:
        parts = str(tanggal_str).lower().split()
        if len(parts) != 3:
            return None
        
        hari = int(parts[0])
        bulan = BULAN_MAP.get(parts[1])
        tahun = int(parts[2])
        
        if not bulan:
            return None
        
        return datetime(tahun, bulan, hari)
    except:
        return None

def clean_number(value):
    """
    Convert Indonesian formatted number to integer
    Example:
    86.937,00 -> 86937
    1.234.567,89 -> 1234567
    """
    if not value:
        return 0

    s = str(value).strip()
    s = s.replace("Rp", "").strip()

    # Ambil bagian sebelum koma (buang desimal)
    if "," in s:
        s = s.split(",", 1)[0]

    # Hapus pemisah ribuan
    s = s.replace(".", "")

    try:
        return int(s)
    except:
        return 0

# ===================== FILE PICKER =====================
pdf_path = filedialog.askopenfilename(
    title="Pilih PDF Faktur Pajak",
    filetypes=[("PDF Files", "*.pdf")]
)

if not pdf_path:
    raise SystemExit("PDF tidak dipilih")

# ===================== READ PDF =====================
reader = PyPDF2.PdfReader(pdf_path)
text = ""
for p in reader.pages:
    text += p.extract_text() + "\n"

lines = [l.strip() for l in text.splitlines() if l.strip()]

# ===================== HELPER =====================
def find_line_start(keyword):
    for i, l in enumerate(lines):
        if l.startswith(keyword):
            return i, l
    return None, None

def find_contains(keyword):
    for i, l in enumerate(lines):
        if keyword in l:
            return i, l
    return None, None

# ===================== HEADER EXTRACTION =====================
data = {}

# Nomor Faktur
_, line = find_contains("Kode dan Nomor Seri Faktur Pajak:")
data["NOMOR_FAKTUR"] = re.findall(r"\d+", line)[-1] if line else ""

# Referensi
_, line = find_contains("(Referensi:")
data["REFERENSI"] = re.findall(r"\(Referensi:\s*([A-Z0-9]+)", line)[0] if line else ""

# Kota & Tanggal
idx, _ = find_contains("secara elektronik sehingga tidak diperlukan tanda tangan basah pada Faktur Pajak ini.")
kota = tanggal = ""
if idx is not None:
    kota_line = lines[idx + 1]
    kota = kota_line.split(",")[0].strip()
    tanggal_raw = " ".join(lines[idx + 1:idx + 3])
    tgl = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", tanggal_raw)
    tanggal = tgl.group(1) if tgl else ""

data["KOTA"] = kota
data["TANGGAL"] = tanggal

# Penandatangan
idx, _ = find_contains("Ditandatangani secara elektronik")
data["PENANDATANGAN"] = lines[idx + 1] if idx else ""

# ===================== SELLER =====================
idx, _ = find_contains("Pengusaha Kena Pajak:")
data["NAMA_PENJUAL"] = lines[idx + 1].replace("Nama :", "").strip()

alamat_penjual = []
for i in range(idx + 2, len(lines)):
    if lines[i].startswith("NPWP"):
        break
    alamat_penjual.append(lines[i].replace("Alamat :", "").strip())
data["ALAMAT_PENJUAL"] = " ".join(alamat_penjual)

# NPWP
npwps = [l.replace("NPWP :", "").strip() for l in lines if l.startswith("NPWP")]
data["NPWP_PENJUAL"] = npwps[0] if len(npwps) > 0 else ""
data["NPWP_PEMBELI"] = npwps[1] if len(npwps) > 1 else ""

# ===================== BUYER =====================
idx, _ = find_contains("Pembeli Barang Kena Pajak")
data["NAMA_PEMBELI"] = lines[idx + 1].replace("Nama :", "").strip()

alamat_pembeli = []
for i in range(idx + 2, len(lines)):
    if lines[i].startswith("NPWP"):
        break
    alamat_pembeli.append(lines[i].replace("Alamat :", "").strip())
alamat = " ".join(alamat_pembeli)
alamat = re.sub(r"#\d+$", "", alamat).strip()
data["ALAMAT_PEMBELI"] = alamat

# Identitas
for k in ["NIK", "Nomor Paspor", "Identitas Lain", "Email"]:
    _, l = find_line_start(k)
    data[k.upper()] = l.split(":", 1)[1].strip() if l else ""

# ===================== TOTAL =====================
for key, prefix in {
    "TOTAL_HARGA": "Harga Jual / Penggantian / Uang Muka / Termin",
    "TOTAL_POTONGAN": "Dikurangi Potongan Harga",
    "DPP": "Dasar Pengenaan Pajak",
    "PPN": "Jumlah PPN (Pajak Pertambahan Nilai) ",
    "TOTAL_PPNBM": "Jumlah PPnBM (Pajak Penjualan atas Barang Mewah) "
}.items():
    _, l = find_line_start(prefix)
    data[key] = l.replace(prefix, "").strip() if l else ""

# ===================== DETAIL BARANG =====================
rows = []
i = 0
faktur_counter = 1  # Counter untuk nomor urut faktur

while i < len(lines):
    if re.match(r"\d+\s+\d{6}", lines[i]):
        kode = re.findall(r"\d{6}", lines[i])[0]
        nama = re.sub(r"^\d+\s+\d{6}", "", lines[i]).strip()

        harga_line = lines[i + 1]
        harga = harga_line.split()[1]
        qty = harga_line.split()[3]
        satuan = harga_line.split()[4]

        potongan = lines[i + 2].split()[-1]

        ppnbm_line = lines[i + 3]
        tarif = re.search(r"\((.*?)\)", ppnbm_line).group(1)
        rp = re.findall(r"Rp\s*([\d.,]+)", ppnbm_line)[0]
        
        if "," in rp:
            after_comma = rp.split(",", 1)[1]
            total = after_comma[2:] if len(after_comma) > 2 else ""
        else:
            total = rp

        rows.append({
            "NO": faktur_counter,  # Kolom NO di paling awal
            **data,
            "KODE_BARANG": kode,
            "NAMA_BARANG": nama,
            "HARGA": harga,
            "QTY": qty,
            "SATUAN": satuan,
            "POTONGAN": potongan,
            "TARIF_PPNBM": tarif,
            "PPNBM": rp[:4],
            "TOTAL_HARGA_BARANG": total
        })
        i += 4
    else:
        i += 1

# Jika ada faktur baru, increment counter (logika bisa disesuaikan)
# Saat ini setiap PDF = 1 faktur, jadi semua rows dapat NO yang sama
# Jika mau auto-detect faktur baru, tambahkan logik di loop

df = pd.DataFrame(rows)

# ===================== SAVE TO EXCEL =====================
save_path = filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel Files", "*.xlsx")]
)

if save_path:
    # Export ke Excel pertama kali
    df.to_excel(save_path, index=False)
    
    # Load kembali untuk formatting
    wb = load_workbook(save_path)
    ws = wb.active
    
    # Map header ke column index
    header_map = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            header_map[str(header).strip()] = col
    
    # Format setiap cell berdasarkan header
    for row_idx in range(2, ws.max_row + 1):
        for header, col_idx in header_map.items():
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            
            # Format TEXT (anti hilang 0 depan)
            if header in TEXT_HEADERS:
                cell.value = str(value) if value else ""
                cell.number_format = "@"
            
            # Format TANGGAL
            elif header in DATE_HEADERS:
                dt = parse_tanggal_teks(value)
                if dt:
                    cell.value = dt
                    cell.number_format = "DD/MM/YYYY"
                else:
                    cell.value = value
            
            # Format NUMBER
            elif header in NUMBER_HEADERS:
                cell.value = clean_number(value)
                cell.number_format = "#,##0"
    
    # ===================== MERGE CELLS =====================
    # Group by NOMOR_FAKTUR untuk tau range merge
    current_faktur = None
    start_row = 2
    
    for row_idx in range(2, ws.max_row + 2):  # +2 untuk handle row terakhir
        if row_idx <= ws.max_row:
            faktur_col = header_map.get("NOMOR_FAKTUR")
            current_val = ws.cell(row=row_idx, column=faktur_col).value if faktur_col else None
        else:
            current_val = None
        
        # Jika faktur berubah atau baris terakhir
        if current_val != current_faktur and current_faktur is not None:
            end_row = row_idx - 1
            
            # Merge kolom header (yang datanya sama) TERMASUK NO
            for header, col_idx in header_map.items():
                if header in MERGE_COLUMNS and end_row > start_row:
                    ws.merge_cells(
                        start_row=start_row,
                        start_column=col_idx,
                        end_row=end_row,
                        end_column=col_idx
                    )
                    # Set alignment ke center-middle untuk merged cells
                    ws.cell(row=start_row, column=col_idx).alignment = Alignment(
                        horizontal='center',
                        vertical='center',
                        wrap_text=True
                    )
            
            start_row = row_idx
        
        current_faktur = current_val
    
    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Auto-adjust column width
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Khusus kolom NO, set width lebih kecil
    if "NO" in header_map:
        no_col_letter = get_column_letter(header_map["NO"])
        ws.column_dimensions[no_col_letter].width = 5
    
    # Border untuk semua cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    
    wb.save(save_path)
    print("✅ FILE BERHASIL DISIMPAN DENGAN KOLOM NO DI PALING AWAL")