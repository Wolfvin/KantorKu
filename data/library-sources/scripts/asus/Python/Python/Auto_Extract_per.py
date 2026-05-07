"""
Smart PDF AutoExtractor - Automated PDF extraction to Excel
Combines e-Faktur and BUPOT extraction capabilities with modern GUI
"""
import sys
import os
from pathlib import Path
import json
import threading
import re
from datetime import datetime
from tkinter import filedialog, messagebox
import io

# Third-party imports
import customtkinter as ctk
import webbrowser
import PyPDF2
import pandas as pd
import fitz  # PyMuPDF for image extraction
import imagehash
from PIL import Image

# ============================================================================================
#                                    SETUP CUSTOMTKINTER
# ============================================================================================
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

# ============================================================================================
#                                    CONFIGURATION
# ============================================================================================
class Config:
    """Application configuration constants"""
    WINDOW_TITLE = "Smart PDF AutoExtractor"
    WINDOW_WIDTH = 1200
    WINDOW_HEIGHT = 750
    MIN_WIDTH = 1000
    MIN_HEIGHT = 650
    CONFIG_FILENAME = "SmartPDFExtractor_config.json"
    
    # UI Colors
    SIDEBAR_COLOR = "#0F172A"
    BUTTON_COLOR = "#1E293B"
    BUTTON_HOVER = "#1E3A8A"
    SUCCESS_COLOR = "#10b981"
    ERROR_COLOR = "#ef4444"
    WARNING_COLOR = "#f59e0b"

# ============================================================================================
#                                    STATUS DETECTION HELPER
# ============================================================================================
def detect_efaktur_status(pdf_path):
    """
    Detect e-Faktur status by hashing first image on page 1
    
    Returns:
        str: "amended", "canceled", "credited", "approved", or "unknown"
    """
    HASH_STATUS_MAP = {
        "c31e32c63de10ee6": "amended",
        "969632c6595b4e6c": "canceled",
        "8000000000000000": "credited",  # For Masukan
        # "8000000000000000": "approved",  # For Keluaran - same hash, different context
    }
    
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        
        images = page.get_images(full=True)
        if not images:
            return "unknown"
        
        # Get first image
        xref = images[0][0]
        pix = fitz.Pixmap(doc, xref)
        
        # Convert to RGB if needed
        if pix.n > 4:
            pix = fitz.Pixmap(fitz.csRGB, pix)
        
        # Create PIL Image
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        
        # Calculate perceptual hash
        h = str(imagehash.phash(img))
        
        # Map hash to status
        status = HASH_STATUS_MAP.get(h, "unknown")
        
        # Special case: hash 8000000000000000 means different things
        # For now return "approved" (can be overridden by context)
        if h == "8000000000000000":
            return "approved"  # Default, can be "credited" for Masukan
        
        doc.close()
        return status
        
    except Exception as e:
        print(f"⚠️ Error detecting status: {e}")
        return "unknown"

# ============================================================================================
#                                    E-FAKTUR EXTRACTION LOGIC
# ============================================================================================
class EFakturExtractor:
    """Handles e-Faktur PDF extraction"""
    
    # Month mapping
    BULAN_MAP = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12
    }
    
    # Text headers (preserve leading zeros)
    TEXT_HEADERS = {
        "NOMOR_FAKTUR", "REFERENSI", "NPWP_PENJUAL", "NPWP_PEMBELI",
        "KODE_BARANG", "NIK", "NOMOR PASPOR", "IDENTITAS LAIN", "EMAIL"
    }
    
    # Date headers
    DATE_HEADERS = {"TANGGAL"}
    
    # Number headers
    NUMBER_HEADERS = {
        "HARGA", "QTY", "POTONGAN", "PPNBM", "TOTAL_HARGA_BARANG",
        "TOTAL_HARGA", "TOTAL_POTONGAN", "DPP", "PPN", "TOTAL_PPNBM", "NO"
    }
    
    # Columns to merge
    MERGE_COLUMNS = {
        "NO", "STATUS", "NOMOR_FAKTUR", "REFERENSI", "KOTA", "TANGGAL", "PENANDATANGAN",
        "NAMA_PENJUAL", "ALAMAT_PENJUAL", "NPWP_PENJUAL", "NPWP_PEMBELI",
        "NAMA_PEMBELI", "ALAMAT_PEMBELI", "NIK", "NOMOR PASPOR",
        "IDENTITAS LAIN", "EMAIL", "TOTAL_HARGA", "TOTAL_POTONGAN",
        "DPP", "PPN", "TOTAL_PPNBM"
    }
    
    @staticmethod
    def parse_tanggal_teks(tanggal_str):
        """Convert '20 Desember 2025' to datetime object"""
        if not tanggal_str or tanggal_str == "":
            return None
        
        try:
            parts = str(tanggal_str).lower().split()
            if len(parts) != 3:
                return None
            
            hari = int(parts[0])
            bulan = EFakturExtractor.BULAN_MAP.get(parts[1])
            tahun = int(parts[2])
            
            if not bulan:
                return None
            
            return datetime(tahun, bulan, hari)
        except:
            return None
    
    @staticmethod
    def clean_number(value):
        """Convert Indonesian formatted number to integer"""
        if not value:
            return 0
        
        s = str(value).strip()
        s = s.replace("Rp", "").strip()
        
        if "," in s:
            s = s.split(",", 1)[0]
        
        s = s.replace(".", "")
        
        try:
            return int(s)
        except:
            return 0
    
    @staticmethod
    def find_line_start(lines, keyword):
        """Find line starting with keyword"""
        for i, l in enumerate(lines):
            if l.startswith(keyword):
                return i, l
        return None, None
    
    @staticmethod
    def find_contains(lines, keyword):
        """Find line containing keyword"""
        for i, l in enumerate(lines):
            if keyword in l:
                return i, l
        return None, None
    
    @classmethod
    def extract(cls, pdf_path, faktur_counter=1, detect_status=True):
        """
        Extract e-Faktur data from PDF
        
        Args:
            pdf_path: Path to PDF file
            faktur_counter: Sequential number for this PDF (1, 2, 3, ...)
            detect_status: Whether to detect status from image hash
        """
        # Detect status from image hash
        status = "unknown"
        if detect_status:
            status = detect_efaktur_status(pdf_path)
        
        # Read PDF
        reader = PyPDF2.PdfReader(pdf_path)
        text = ""
        for p in reader.pages:
            text += p.extract_text() + "\n"
        
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        
        # Extract header data
        data = {}
        
        # Add STATUS as first field
        data["STATUS"] = status
        
        # Nomor Faktur
        _, line = cls.find_contains(lines, "Kode dan Nomor Seri Faktur Pajak:")
        data["NOMOR_FAKTUR"] = re.findall(r"\d+", line)[-1] if line else ""
        
        # Referensi
        _, line = cls.find_contains(lines, "(Referensi:")
        data["REFERENSI"] = re.findall(r"\(Referensi:\s*([A-Z0-9]+)", line)[0] if line else ""
        
        # Kota & Tanggal
        idx, _ = cls.find_contains(lines, "secara elektronik sehingga tidak diperlukan tanda tangan basah pada Faktur Pajak ini.")
        kota = tanggal = ""
        if idx is not None:
            kota_line = lines[idx + 1]
            kota = kota_line.split(",")[0].strip()
            tanggal_raw = " ".join(lines[idx + 1:idx + 3])
            tgl = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", tanggal_raw)
            tanggal = tgl.group(1) if tgl else ""
        
        data["KOTA"] = kota
        data["TANGGAL"] = tanggal
        
        # Penandatangan
        idx, _ = cls.find_contains(lines, "Ditandatangani secara elektronik")
        data["PENANDATANGAN"] = lines[idx + 1] if idx else ""
        
        # Seller
        idx, _ = cls.find_contains(lines, "Pengusaha Kena Pajak:")
        data["NAMA_PENJUAL"] = lines[idx + 1].replace("Nama :", "").strip()
        
        alamat_penjual = []
        for i in range(idx + 2, len(lines)):
            if lines[i].startswith("NPWP"):
                break
            alamat_penjual.append(lines[i].replace("Alamat :", "").strip())
        data["ALAMAT_PENJUAL"] = " ".join(alamat_penjual)
        
        # NPWP
        npwps = [l.replace("NPWP :", "").strip() for l in lines if l.startswith("NPWP")]
        data["NPWP_PENJUAL"] = npwps[0] if len(npwps) > 0 else ""
        data["NPWP_PEMBELI"] = npwps[1] if len(npwps) > 1 else ""
        
        # Buyer
        idx, _ = cls.find_contains(lines, "Pembeli Barang Kena Pajak")
        data["NAMA_PEMBELI"] = lines[idx + 1].replace("Nama :", "").strip()
        
        alamat_pembeli = []
        for i in range(idx + 2, len(lines)):
            if lines[i].startswith("NPWP"):
                break
            alamat_pembeli.append(lines[i].replace("Alamat :", "").strip())
        alamat = " ".join(alamat_pembeli)
        alamat = re.sub(r"#\d+$", "", alamat).strip()
        data["ALAMAT_PEMBELI"] = alamat
        
        # Identity
        for k in ["NIK", "Nomor Paspor", "Identitas Lain", "Email"]:
            _, l = cls.find_line_start(lines, k)
            data[k.upper()] = l.split(":", 1)[1].strip() if l else ""
        
        # Totals
        for key, prefix in {
            "TOTAL_HARGA": "Harga Jual / Penggantian / Uang Muka / Termin",
            "TOTAL_POTONGAN": "Dikurangi Potongan Harga",
            "DPP": "Dasar Pengenaan Pajak",
            "PPN": "Jumlah PPN (Pajak Pertambahan Nilai) ",
            "TOTAL_PPNBM": "Jumlah PPnBM (Pajak Penjualan atas Barang Mewah) "
        }.items():
            _, l = cls.find_line_start(lines, prefix)
            data[key] = l.replace(prefix, "").strip() if l else ""
        
        # Detail barang - SEMUA rows dari PDF ini dapat nomor yang SAMA
        rows = []
        i = 0
        
        while i < len(lines):
            if re.match(r"\d+\s+\d{6}", lines[i]):
                kode = re.findall(r"\d{6}", lines[i])[0]
                nama = re.sub(r"^\d+\s+\d{6}", "", lines[i]).strip()
                
                harga_line = lines[i + 1]
                harga = harga_line.split()[1]
                qty = harga_line.split()[3]
                satuan = harga_line.split()[4]
                
                potongan = lines[i + 2].split()[-1]
                
                ppnbm_line = lines[i + 3]
                tarif = re.search(r"\((.*?)\)", ppnbm_line).group(1)
                rp = re.findall(r"Rp\s*([\d.,]+)", ppnbm_line)[0]
                
                if "," in rp:
                    after_comma = rp.split(",", 1)[1]
                    total = after_comma[2:] if len(after_comma) > 2 else ""
                else:
                    total = rp
                
                rows.append({
                    "NO": faktur_counter,  # Semua barang dalam 1 PDF = nomor sama
                    **data,
                    "KODE_BARANG": kode,
                    "NAMA_BARANG": nama,
                    "HARGA": harga,
                    "QTY": qty,
                    "SATUAN": satuan,
                    "POTONGAN": potongan,
                    "TARIF_PPNBM": tarif,
                    "PPNBM": rp[:4],
                    "TOTAL_HARGA_BARANG": total
                })
                i += 4
            else:
                i += 1
        
        return pd.DataFrame(rows)

# ============================================================================================
#                                    BUPOT EXTRACTION LOGIC
# ============================================================================================
class BUPOTExtractor:
    """Handles BUPOT PDF extraction"""
    
    BULAN_MAP = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12
    }
    
    @staticmethod
    def ubah_tanggal(teks_tanggal):
        """Convert date text to datetime"""
        if not teks_tanggal:
            return ""
        teks_tanggal = teks_tanggal.strip().lower()
        match = re.match(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})", teks_tanggal)
        if match:
            hari = int(match.group(1))
            bulan = BUPOTExtractor.BULAN_MAP.get(match.group(2))
            tahun = int(match.group(3))
            if bulan:
                return datetime(tahun, bulan, hari)
        match = re.match(r"(\d{2})-(\d{4})", teks_tanggal)
        if match:
            bulan = int(match.group(1))
            tahun = int(match.group(2))
            return datetime(tahun, bulan, 1)
        return ""
    
    @staticmethod
    def bersihkan_angka(teks_angka):
        """Clean number text"""
        if not teks_angka:
            return ""
        angka = re.sub(r"[^\d]", "", teks_angka)
        return int(angka) if angka else ""
    
    @staticmethod
    def deteksi_jenis_dokumen(teks):
        """Detect document type (BPPU or BP21)"""
        lines = teks.splitlines()
        for i, baris in enumerate(lines):
            if "UNIFIKASI BERFORMAT STANDAR" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip().replace(" ", "")
                    if next_line == "BPPU":
                        return "BPPU"
            if "PASAL 21 YANG BERSIFAT FINAL" in baris:
                kata_terakhir = baris.strip().split()[-1]
                if kata_terakhir == "BP21":
                    return "BP21"
        return None
    
    @classmethod
    def ekstrak_bppu(cls, teks):
        """Extract BPPU data"""
        lines = teks.splitlines()
        
        # Nomor Bukti Potong, Masa Pajak, Sifat, Status
        nomor_bukti = masa_pajak = sifat_bukti = status_bukti = ""
        for i, baris in enumerate(lines):
            if "A. IDENTITAS WAJIB PAJAK YANG DIPOTONG DAN/ATAU DIPUNGUT PPh ATAU PENERIMA PENGHASILAN" in baris:
                if i > 0:
                    prev_line = lines[i - 1].strip()
                    parts = prev_line.split()
                    if len(parts) >= 4:
                        nomor_bukti = parts[0]
                        masa_pajak = parts[1]
                        sifat_bukti = " ".join(parts[2:-1])
                        status_bukti = parts[-1]
                break
        
        masa_pajak_obj = cls.ubah_tanggal(masa_pajak)
        
        # Kode Objek Pajak
        kode_objek = ""
        for i, baris in enumerate(lines):
            if "B.3 B.4 B.5 B.6 B.7" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    kode_objek = next_line.split()[0]
                break
        
        # DPP dan PPH
        dpp = pph = ""
        for i, baris in enumerate(lines):
            if "B.8 Dokumen Dasar Bukti" in baris:
                if i >= 2:
                    target_line = lines[i - 2].strip()
                    angka_semua = re.findall(r'\d[\d\.]*', target_line)
                    if len(angka_semua) >= 3:
                        dpp = angka_semua[-3]
                        pph = angka_semua[-1]
                break
        
        # NPWP Pemotong
        npwp_pemotong = ""
        for baris in lines:
            if "C.1 NPWP / NIK :" in baris:
                npwp_pemotong = baris.split(":")[-1].strip()
                break
        
        # Nama Pemotong
        nama_pemotong = ""
        for i, baris in enumerate(lines):
            if "C.3 NAMA PEMOTONG DAN/ATAU PEMUNGUT" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    if "PPh:" in next_line:
                        nama_pemotong = next_line.split(":")[-1].strip()
                break
        
        # Tanggal Bukti Potong
        tanggal_bukti = ""
        for baris in lines:
            if "C.4 TANGGAL :" in baris:
                tanggal_bukti = baris.split(":")[-1].strip()
                break
        
        tanggal_bukti_obj = cls.ubah_tanggal(tanggal_bukti)
        
        return {
            'JENIS DOKUMEN': 'BPPU',
            'NOMOR BUKTI POTONG': nomor_bukti,
            'MASA PAJAK': masa_pajak_obj,
            'SIFAT BUKTI POTONG': sifat_bukti,
            'STATUS BUKTI POTONG': status_bukti,
            'KODE OBJEK PAJAK': kode_objek,
            'DPP': cls.bersihkan_angka(dpp),
            'PPH YANG DIPOTONG': cls.bersihkan_angka(pph),
            'NPWP PEMOTONG': npwp_pemotong,
            'NAMA PEMOTONG': nama_pemotong,
            'TANGGAL BUKTI POTONG': tanggal_bukti_obj
        }
    
    @classmethod
    def ekstrak_bp21(cls, teks):
        """Extract BP21 data"""
        lines = teks.splitlines()
        
        # Nomor Bukti Potong, Masa Pajak, Sifat, Status
        nomor_bukti = masa_pajak = sifat_bukti = status_bukti = ""
        for i, baris in enumerate(lines):
            if "NOMOR BUKTI PEMOTONGAN MASA PAJAK SIFAT PEMOTONGAN STATUS BUKTI PEMOTONGAN" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    parts = next_line.split()
                    if len(parts) >= 4:
                        nomor_bukti = parts[0]
                        masa_pajak = parts[1]
                        sifat_bukti = " ".join(parts[2:-1])
                        status_bukti = parts[-1]
                break
        
        masa_pajak_obj = cls.ubah_tanggal(masa_pajak)
        
        # Kode Objek Pajak
        kode_objek = ""
        for i, baris in enumerate(lines):
            if "B.2 B.3 B.4 B.5 B.6 B.7" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    kode_objek = next_line.split()[0]
                break
        
        # DPP dan PPH
        dpp = pph = ""
        for i, baris in enumerate(lines):
            if "B.8 Dokumen Referensi Jenis Dokumen" in baris or "B.8 Dokumen Referensi" in baris:
                if i >= 1:
                    target_line = lines[i - 1].strip()
                    angka_bersih = re.sub(r'[A-Za-z\s]+', ' ', target_line)
                    angka_semua = angka_bersih.split()
                    if len(angka_semua) >= 4:
                        dpp = angka_semua[0]
                        pph = angka_semua[-1]
                break
        
        # NPWP Pemotong
        npwp_pemotong = ""
        for baris in lines:
            if "C.1 NPWP/NIK :" in baris:
                npwp_pemotong = baris.split(":")[-1].strip()
                break
        
        # Nama Pemotong
        nama_pemotong = ""
        for baris in lines:
            if "C.3 Nama Pemotong :" in baris:
                nama_pemotong = baris.split(":")[-1].strip()
                break
        
        # Tanggal Bukti Potong
        tanggal_bukti = ""
        for baris in lines:
            if "C.4 Tanggal :" in baris:
                tanggal_bukti = baris.split(":")[-1].strip()
                break
        
        tanggal_bukti_obj = cls.ubah_tanggal(tanggal_bukti)
        
        return {
            'JENIS DOKUMEN': 'BP21',
            'NOMOR BUKTI POTONG': nomor_bukti,
            'MASA PAJAK': masa_pajak_obj,
            'SIFAT BUKTI POTONG': sifat_bukti,
            'STATUS BUKTI POTONG': status_bukti,
            'KODE OBJEK PAJAK': kode_objek,
            'DPP': cls.bersihkan_angka(dpp),
            'PPH YANG DIPOTONG': cls.bersihkan_angka(pph),
            'NPWP PEMOTONG': npwp_pemotong,
            'NAMA PEMOTONG': nama_pemotong,
            'TANGGAL BUKTI POTONG': tanggal_bukti_obj
        }
    
    @classmethod
    def extract(cls, pdf_path):
        """Extract BUPOT data from PDF"""
        reader = PyPDF2.PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            pdf_text += page.extract_text() or ""
        
        jenis_dokumen = cls.deteksi_jenis_dokumen(pdf_text)
        
        if jenis_dokumen == "BPPU":
            data = cls.ekstrak_bppu(pdf_text)
        elif jenis_dokumen == "BP21":
            data = cls.ekstrak_bp21(pdf_text)
        else:
            raise ValueError("Jenis dokumen tidak dikenali")
        
        return pd.DataFrame([data])

# ============================================================================================
#                                    MAIN APPLICATION
# ============================================================================================
class SmartPDFExtractor(ctk.CTk):
    
    def __init__(self):
        # Expiry check
        if datetime.now() > datetime(2026, 5, 2, 23, 59, 59):
            return
        
        super().__init__()
        
        # Window setup
        self.title(Config.WINDOW_TITLE)
        self.geometry(f"{Config.WINDOW_WIDTH}x{Config.WINDOW_HEIGHT}")
        self.minsize(Config.MIN_WIDTH, Config.MIN_HEIGHT)
        
        # Load icon
        try:
            if hasattr(sys, "_MEIPASS"):
                assets_folder = Path(sys._MEIPASS) / "assets"
            else:
                assets_folder = Path("assets")
            
            icon_files = list(assets_folder.glob("*.ico"))
            if icon_files:
                self.iconbitmap(str(icon_files[0]))
        except:
            pass
        
        # Initialize variables
        self.current_page_index = 0
        self.selected_files = []
        self.final_df_global = None
        
        # Config
        self.config_file = self._get_config_path()
        self.load_config()
        
        # Setup UI
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)
        
        # Create UI
        self.create_sidebar()
        
        # Main pages frame
        self.pages_frame = ctk.CTkFrame(self, fg_color="#111827")
        self.pages_frame.grid(row=0, column=1, sticky="nsew")
        self.pages_frame.grid_rowconfigure(0, weight=1)
        self.pages_frame.grid_columnconfigure(0, weight=1)
        
        # Create pages
        self.pages = []
        
        # Home page
        self.home_page = ctk.CTkScrollableFrame(self.pages_frame, fg_color="#1E293B", corner_radius=15)
        self.home_page.grid(row=0, column=0, sticky="nsew")
        self.build_home_page()
        self.pages.append(self.home_page)
        
        # E-Faktur Keluaran page
        self.efaktur_keluaran_page = ctk.CTkScrollableFrame(self.pages_frame, fg_color="#1E293B", corner_radius=15)
        self.efaktur_keluaran_page.grid(row=0, column=0, sticky="nsew")
        self.build_extractor_page(self.efaktur_keluaran_page, "E-Faktur Keluaran Extractor", "efaktur_keluaran")
        self.pages.append(self.efaktur_keluaran_page)
        
        # E-Faktur Masukan page
        self.efaktur_masukan_page = ctk.CTkScrollableFrame(self.pages_frame, fg_color="#1E293B", corner_radius=15)
        self.efaktur_masukan_page.grid(row=0, column=0, sticky="nsew")
        self.build_extractor_page(self.efaktur_masukan_page, "E-Faktur Masukan Extractor", "efaktur_masukan")
        self.pages.append(self.efaktur_masukan_page)
        
        # BUPOT page
        self.bupot_page = ctk.CTkScrollableFrame(self.pages_frame, fg_color="#1E293B", corner_radius=15)
        self.bupot_page.grid(row=0, column=0, sticky="nsew")
        self.build_extractor_page(self.bupot_page, "BUPOT Extractor", "bupot")
        self.pages.append(self.bupot_page)
        
        # Show home
        self.show_page(0)
    
    # ============================================================================================
    #                                    SIDEBAR
    # ============================================================================================
    def create_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=80, fg_color=Config.SIDEBAR_COLOR)
        self.sidebar.grid(row=0, column=0, sticky="ns")
        self.sidebar.grid_rowconfigure(10, weight=1)
        self.sidebar.pack_propagate(False)
        
        self.btn_home = ctk.CTkButton(
            self.sidebar, text="🏠", font=ctk.CTkFont(size=30),
            width=60, height=60, corner_radius=12,
            fg_color=Config.BUTTON_COLOR, hover_color=Config.BUTTON_HOVER,
            command=lambda: self.show_page(0)
        )
        self.btn_home.pack(pady=(20, 10))
        
        # E-Faktur Keluaran
        self.btn_efaktur_keluaran = ctk.CTkButton(
            self.sidebar, text="📤", font=ctk.CTkFont(size=30),
            width=60, height=60, corner_radius=12,
            fg_color=Config.BUTTON_COLOR, hover_color=Config.BUTTON_HOVER,
            command=lambda: self.show_page(1)
        )
        self.btn_efaktur_keluaran.pack(pady=10)
        
        # E-Faktur Masukan
        self.btn_efaktur_masukan = ctk.CTkButton(
            self.sidebar, text="📥", font=ctk.CTkFont(size=30),
            width=60, height=60, corner_radius=12,
            fg_color=Config.BUTTON_COLOR, hover_color=Config.BUTTON_HOVER,
            command=lambda: self.show_page(2)
        )
        self.btn_efaktur_masukan.pack(pady=10)
        
        # BUPOT
        self.btn_bupot = ctk.CTkButton(
            self.sidebar, text="📃", font=ctk.CTkFont(size=30),
            width=60, height=60, corner_radius=12,
            fg_color=Config.BUTTON_COLOR, hover_color=Config.BUTTON_HOVER,
            command=lambda: self.show_page(3)
        )
        self.btn_bupot.pack(pady=10)
    
    # ============================================================================================
    #                                    HOME PAGE
    # ============================================================================================
    def build_home_page(self):
        page = self.home_page
        
        # Logo & Title
        logo_frame = ctk.CTkFrame(page, fg_color="#22D3EE", corner_radius=20)
        logo_frame.pack(pady=20, padx=40, fill="x")
        
        ctk.CTkLabel(
            logo_frame, 
            text="⬢ Smart PDF AutoExtractor", 
            font=ctk.CTkFont(size=28, weight="bold"), 
            text_color="#0F172A"
        ).pack(pady=15)
        
        ctk.CTkLabel(
            logo_frame, 
            text="Ekstraksi Otomatis PDF ke Excel untuk E-Faktur & BUPOT", 
            font=ctk.CTkFont(size=16), 
            text_color="#0F172A"
        ).pack(pady=(0, 15))
        
        # Instructions
        ctk.CTkLabel(
            page, 
            text="Cara Penggunaan", 
            font=ctk.CTkFont(size=22, weight="bold")
        ).pack(pady=(10, 5), anchor="w", padx=40)
        
        instructions = [
            "1. Pilih jenis ekstraksi (E-Faktur atau BUPOT) dari menu di kiri",
            "2. Upload file PDF yang ingin diekstrak (bisa multiple files)",
            "3. Klik 'Proses Sekarang' untuk memulai ekstraksi",
            "4. Download hasil dalam format Excel dengan formatting profesional",
            "5. File Excel sudah include cell merging, number formatting, borders, etc",
        ]
        
        for ins in instructions:
            ctk.CTkLabel(page, text=ins, anchor="w").pack(fill="x", padx=60)
        
        # Features
        ctk.CTkLabel(
            page, 
            text="Fitur Utama", 
            font=ctk.CTkFont(size=22, weight="bold")
        ).pack(pady=(15, 5), anchor="w", padx=40)
        
        features = [
            "✅ E-Faktur: Ekstrak nomor, referensi, NPWP, nama, alamat, detail barang",
            "✅ BUPOT: Auto-detect BPPU atau BP21, ekstrak lengkap data pajak",
            "✅ Format Excel profesional dengan borders, alignment, merged cells",
            "✅ Validasi data otomatis untuk deteksi cell kosong",
            "✅ Support multiple PDF dalam satu batch",
            "✅ Number formatting (ribuan, tanggal) otomatis",
        ]
        
        for feat in features:
            ctk.CTkLabel(page, text=feat, anchor="w").pack(fill="x", padx=60)
        
        # Promo
        promo_frame = ctk.CTkFrame(
            page, 
            fg_color="#111827", 
            corner_radius=20, 
            border_width=2, 
            border_color="#F59E0B"
        )
        promo_frame.pack(padx=40, pady=20, fill="x")
        
        ctk.CTkLabel(
            promo_frame, 
            text="Sambil nungguin laper yaaa? Coba snack di 🥰✨Kedai Ayam Warisan 81✨🥰", 
            justify="center"
        ).pack(pady=10)
        
        btn_frame = ctk.CTkFrame(promo_frame, fg_color="transparent")
        btn_frame.pack()
        
        ctk.CTkButton(
            btn_frame, 
            text="🛵 GoFood", 
            fg_color="#EF4444", 
            hover_color="#F87171",
            command=lambda: webbrowser.open("https://gofood.link/a/Mtv3P3L")
        ).pack(side="left", padx=10, pady=10)
        
        ctk.CTkButton(
            btn_frame, 
            text="🟢 GrabFood", 
            fg_color="#10B981", 
            hover_color="#34D399",
            command=lambda: webbrowser.open("https://r.grab.com/g/6-20260202_212151_1f6a2784162b40d5bd1c465b0e817b13_MEXMPS-6-C6NAVVMAVTDEFE")
        ).pack(side="left", padx=10, pady=10)
        
        # WA
        wa_label = ctk.CTkLabel(
            promo_frame, 
            text="📞 089671139111 (Klik untuk menyalin)", 
            text_color="#22D3EE", 
            cursor="hand2"
        )
        wa_label.pack(pady=(5, 10))
        wa_label.bind("<Button-1>", lambda e: self.copy_to_clipboard("089671139111"))
        
        # Credit
        ctk.CTkLabel(
            page, 
            text="BY RAYMOND FO", 
            text_color="gray"
        ).pack(anchor="e", pady=(0, 5), padx=40)
    
    # ============================================================================================
    #                                    EXTRACTOR PAGE
    # ============================================================================================
    def build_extractor_page(self, page, title_text, mode):
        """Build extractor page"""
        
        # Header
        header_frame = ctk.CTkFrame(
            page, 
            fg_color="#111827", 
            corner_radius=15,
            border_width=1, 
            border_color="#1F2937"
        )
        header_frame.pack(fill="x", padx=20, pady=(15, 10))
        
        top_bar = ctk.CTkFrame(header_frame, height=4, fg_color="#3B82F6")
        top_bar.pack(fill="x")
        
        title_label = ctk.CTkLabel(
            header_frame,
            text="🚀 " + title_text,
            font=ctk.CTkFont(size=25, weight="bold"),
            text_color="#F1F5F9"
        )
        title_label.pack(anchor="w", padx=15, pady=(10, 2))
        
        subtitle = ctk.CTkLabel(
            header_frame,
            text="Upload PDF dan ekstrak data ke Excel otomatis",
            font=ctk.CTkFont(size=12),
            text_color="#94A3B8"
        )
        subtitle.pack(anchor="w", padx=15, pady=(0, 10))
        
        # Upload Frame
        upload_frame = ctk.CTkFrame(
            page, 
            fg_color="#111827", 
            corner_radius=20, 
            border_width=2, 
            border_color="#22D3EE"
        )
        upload_frame.pack(padx=40, pady=10, fill="x")
        
        ctk.CTkLabel(
            upload_frame, 
            text="Upload File PDF"
        ).pack(pady=(15, 5))
        
        def pilih_file():
            files = filedialog.askopenfilenames(filetypes=[("PDF Files", "*.pdf")])
            if files:
                self.selected_files = list(files)
                tampilkan_file()
                status_label.configure(text=f"{len(self.selected_files)} file siap diproses")
        
        ctk.CTkButton(
            upload_frame,
            text="PILIH FILE PDF",
            fg_color="#22D3EE",
            hover_color="#67E8F9",
            text_color="black",
            corner_radius=30,
            height=40,
            command=pilih_file
        ).pack(pady=10)
        
        # File List
        file_frame = ctk.CTkFrame(page, fg_color="#111827", corner_radius=20)
        file_frame.pack(padx=40, pady=10, fill="both", expand=True)
        
        file_listbox = ctk.CTkTextbox(file_frame, height=150)
        file_listbox.pack(padx=15, pady=15, fill="both", expand=True)
        
        def tampilkan_file():
            file_listbox.delete("1.0", "end")
            for f in self.selected_files:
                file_listbox.insert("end", f"✔ {os.path.basename(f)}\n")
        
        # Process Button
        ctk.CTkButton(
            page,
            text="⚡ PROSES SEKARANG",
            fg_color="#22C55E",
            hover_color="#4ADE80",
            corner_radius=30,
            height=45,
            command=lambda: self.process_extraction(mode)
        ).pack(pady=12)
        
        # Status
        status_label = ctk.CTkLabel(page, text="Status: Menunggu file...", text_color="gray")
        status_label.pack()
        
        # Download Button
        ctk.CTkButton(
            page,
            text="⬇ DOWNLOAD HASIL EXCEL",
            fg_color="#A78BFA",
            hover_color="#C4B5FD",
            corner_radius=30,
            height=45,
            command=lambda: self.download_excel(mode)
        ).pack(pady=15)
        
        # Store references
        if mode == "efaktur_keluaran":
            self.efaktur_keluaran_status = status_label
            self.efaktur_keluaran_filelist = file_listbox
        elif mode == "efaktur_masukan":
            self.efaktur_masukan_status = status_label
            self.efaktur_masukan_filelist = file_listbox
        else:  # bupot
            self.bupot_status = status_label
            self.bupot_filelist = file_listbox
    
    # ============================================================================================
    #                                    PROCESSING
    # ============================================================================================
    def process_extraction(self, mode):
        """Process extraction"""
        if not self.selected_files:
            messagebox.showwarning("Peringatan", "Pilih file PDF dulu!")
            return
        
        threading.Thread(
            target=self._extract_pdfs,
            args=(mode,),
            daemon=True
        ).start()
    
    def _extract_pdfs(self, mode):
        """Extract PDFs in background"""
        if mode == "efaktur_keluaran":
            status_label = self.efaktur_keluaran_status
        elif mode == "efaktur_masukan":
            status_label = self.efaktur_masukan_status
        else:
            status_label = self.bupot_status
        
        self.after(0, lambda: status_label.configure(text="🔄 Memproses dokumen..."))
        
        dfs = []
        gagal_files = []
        
        for idx, f in enumerate(self.selected_files, start=1):  # Start from 1 for counter
            try:
                if mode in ["efaktur_keluaran", "efaktur_masukan"]:
                    # Both use same extractor, just different context for hash interpretation
                    df = EFakturExtractor.extract(f, faktur_counter=idx, detect_status=True)
                    
                    # For Masukan, override "approved" to "credited" if hash is 8000000000000000
                    if mode == "efaktur_masukan":
                        df.loc[df['STATUS'] == 'approved', 'STATUS'] = 'credited'
                else:
                    df = BUPOTExtractor.extract(f)
                
                # Validate
                kosong_rows = df[df.apply(lambda x: x.isnull() | x.astype(str).str.strip().eq('')).any(axis=1)]
                if not kosong_rows.empty:
                    baris_error = (kosong_rows.index + 2).tolist()
                    gagal_files.append((os.path.basename(f), f"Baris {baris_error} kosong"))
                
                dfs.append(df)
                    
            except Exception as e:
                gagal_files.append((os.path.basename(f), str(e)))
        
        if dfs:
            self.final_df_global = pd.concat(dfs, ignore_index=True)
            
            berhasil = len(self.selected_files) - len(gagal_files)
            pesan = f"✅ {berhasil} PDF berhasil diproses."
            
            if gagal_files:
                pesan += "\n\n⚠ Beberapa file gagal:\n"
                for fname, error in gagal_files:
                    pesan += f"- {fname}: {error}\n"
            
            self.after(0, lambda: status_label.configure(
                text=f"✅ Selesai! {berhasil} berhasil, {len(gagal_files)} gagal"
            ))
            self.after(0, lambda: messagebox.showinfo("Hasil Ekstraksi", pesan))
        else:
            self.after(0, lambda: status_label.configure(text="❌ Tidak ada PDF yang berhasil!"))
            self.after(0, lambda: messagebox.showerror("Error", "Tidak ada PDF yang berhasil diekstrak!"))
    
    def download_excel(self, mode):
        """Download Excel file with complete formatting and merging logic"""
        if self.final_df_global is None:
            messagebox.showerror("Error", "Data belum diproses!")
            return
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        
        if not save_path:
            return
        
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Save initial Excel
        self.final_df_global.to_excel(save_path, index=False)
        
        # Reload for formatting
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
        
        if mode in ["efaktur_keluaran", "efaktur_masukan"]:
            # ==================== E-FAKTUR FORMATTING ====================
            
            # Map header to column index
            header_map = {}
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    header_map[str(header).strip()] = col
            
            # Format each cell based on header type
            for row_idx in range(2, ws.max_row + 1):
                for header, col_idx in header_map.items():
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    # Format TEXT (preserve leading zeros)
                    if header in EFakturExtractor.TEXT_HEADERS:
                        cell.value = str(value) if value else ""
                        cell.number_format = "@"
                    
                    # Format DATE
                    elif header in EFakturExtractor.DATE_HEADERS:
                        dt = EFakturExtractor.parse_tanggal_teks(value)
                        if dt:
                            cell.value = dt
                            cell.number_format = "DD/MM/YYYY"
                        else:
                            cell.value = value
                    
                    # Format NUMBER
                    elif header in EFakturExtractor.NUMBER_HEADERS:
                        cell.value = EFakturExtractor.clean_number(value)
                        cell.number_format = "#,##0"
            
            # ==================== MERGE CELLS LOGIC ====================
            # Group by NOMOR_FAKTUR to find merge ranges
            current_faktur = None
            start_row = 2
            
            for row_idx in range(2, ws.max_row + 2):  # +2 to handle last row
                if row_idx <= ws.max_row:
                    faktur_col = header_map.get("NOMOR_FAKTUR")
                    current_val = ws.cell(row=row_idx, column=faktur_col).value if faktur_col else None
                else:
                    current_val = None
                
                # If faktur changes or last row
                if current_val != current_faktur and current_faktur is not None:
                    end_row = row_idx - 1
                    
                    # Merge header columns (including NO & STATUS) for this faktur's transactions
                    for header, col_idx in header_map.items():
                        if header in EFakturExtractor.MERGE_COLUMNS and end_row > start_row:
                            ws.merge_cells(
                                start_row=start_row,
                                start_column=col_idx,
                                end_row=end_row,
                                end_column=col_idx
                            )
                            # Center alignment for merged cells
                            ws.cell(row=start_row, column=col_idx).alignment = Alignment(
                                horizontal='center',
                                vertical='center',
                                wrap_text=True
                            )
                    
                    start_row = row_idx
                
                current_faktur = current_val
            
            # Set NO column width smaller
            if "NO" in header_map:
                no_col_letter = get_column_letter(header_map["NO"])
                ws.column_dimensions[no_col_letter].width = 5
            
            # Set STATUS column width
            if "STATUS" in header_map:
                status_col_letter = get_column_letter(header_map["STATUS"])
                ws.column_dimensions[status_col_letter].width = 12
        
        else:
            # ==================== BUPOT FORMATTING ====================
            # Format specific columns
            for cell in ws['I']:  # NPWP PEMOTONG
                if cell.row > 1:
                    cell.number_format = '@'
            
            for cell in ws['G']:  # DPP
                if cell.row > 1:
                    cell.number_format = '#,##0'
            
            for cell in ws['H']:  # PPH YANG DIPOTONG
                if cell.row > 1:
                    cell.number_format = '#,##0'
            
            for cell in ws['C']:  # MASA PAJAK
                if cell.row > 1:
                    cell.number_format = 'DD/MM/YYYY'
            
            for cell in ws['K']:  # TANGGAL BUKTI POTONG
                if cell.row > 1:
                    cell.number_format = 'DD/MM/YYYY'
        
        # ==================== COMMON FORMATTING ====================
        
        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Borders for all cells
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
        
        wb.save(save_path)
        messagebox.showinfo("Berhasil", f"✅ File Excel berhasil disimpan dengan formatting lengkap!\n\n{save_path}")
    
    # ============================================================================================
    #                                    UTILITY
    # ============================================================================================
    def show_page(self, index):
        """Show page"""
        self.current_page_index = index
        self.highlight_menu(index)
        
        for i, page in enumerate(self.pages):
            if i == index:
                page.lift()
    
    def highlight_menu(self, index):
        """Highlight menu"""
        default = Config.BUTTON_COLOR
        hover = Config.BUTTON_HOVER
        active = "#2563EB"
        active_h = "#1D4ED8"
        
        self.btn_home.configure(fg_color=default, hover_color=hover)
        self.btn_efaktur_keluaran.configure(fg_color=default, hover_color=hover)
        self.btn_efaktur_masukan.configure(fg_color=default, hover_color=hover)
        self.btn_bupot.configure(fg_color=default, hover_color=hover)
        
        if index == 0:
            self.btn_home.configure(fg_color=active, hover_color=active_h)
        elif index == 1:
            self.btn_efaktur_keluaran.configure(fg_color=active, hover_color=active_h)
        elif index == 2:
            self.btn_efaktur_masukan.configure(fg_color=active, hover_color=active_h)
        elif index == 3:
            self.btn_bupot.configure(fg_color=active, hover_color=active_h)
    
    def copy_to_clipboard(self, text):
        """Copy to clipboard"""
        self.clipboard_clear()
        self.clipboard_append(text)
        messagebox.showinfo("Disalin", "Nomor berhasil disalin!")
    
    def _get_config_path(self):
        """Get config path"""
        if os.name == 'nt':
            base = Path(os.getenv('APPDATA', ''))
        else:
            base = Path.home() / '.config'
        
        config_dir = base / 'SmartPDFExtractor'
        config_dir.mkdir(parents=True, exist_ok=True)
        
        return config_dir / Config.CONFIG_FILENAME
    
    def load_config(self):
        """Load config"""
        if self.config_file.exists():
            try:
                with open(self.config_file, "r") as f:
                    data = json.load(f)
            except:
                pass
    
    def save_config(self):
        """Save config"""
        data = {}
        try:
            with open(self.config_file, "w") as f:
                json.dump(data, f, indent=2)
        except:
            pass

# ============================================================================================
#                                    MAIN
# ============================================================================================
if __name__ == "__main__":
    app = SmartPDFExtractor()
    app.mainloop()