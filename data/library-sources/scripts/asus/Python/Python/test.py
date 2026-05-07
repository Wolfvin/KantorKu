"""
Smart PDF AutoExtractor - PySide6 Version
COMPLETE EDITION with Full Drag & Drop Support
"""
import sys
import os
from pathlib import Path
import json
import threading
import re
from datetime import datetime
import string

# PySide6 imports
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QScrollArea, QFrame, QGridLayout,
    QTextEdit, QFileDialog, QMessageBox, QSizePolicy
)
from PySide6.QtCore import Qt, Signal, QSize, QTimer, QMimeData, QPoint
from PySide6.QtGui import QFont, QColor, QPalette, QCursor, QDrag, QPixmap

# PDF and data processing
import PyPDF2
import pandas as pd
import fitz  # PyMuPDF
import imagehash
from PIL import Image
import io
import webbrowser

# ============================================================================================
#                                    CONFIGURATION
# ============================================================================================
class Config:
    """Application configuration constants"""
    WINDOW_TITLE = "Smart PDF AutoExtractor"
    WINDOW_WIDTH = 1200
    WINDOW_HEIGHT = 750
    MIN_WIDTH = 1000
    MIN_HEIGHT = 650
    CONFIG_FILENAME = "SmartPDFExtractor_config.json"
    
    # Colors (Qt compatible)
    COLOR_BG_DARK = "#111827"
    COLOR_BG_MEDIUM = "#1E293B"
    COLOR_SIDEBAR = "#0F172A"
    COLOR_BUTTON = "#1E3A8A"
    COLOR_BUTTON_HOVER = "#2563EB"
    COLOR_BUTTON_ACTIVE = "#3B82F6"
    COLOR_SUCCESS = "#10b981"
    COLOR_ERROR = "#ef4444"
    COLOR_WARNING = "#f59e0b"
    COLOR_ACCENT = "#22D3EE"
    
    # Mapping Colors
    COLOR_FIELD = "#1E3A8A"
    COLOR_FIELD_HOVER = "#2563EB"
    COLOR_FIELD_SELECTED = "#FBBF24"
    COLOR_SLOT_EMPTY = "#374151"
    COLOR_SLOT_FILLED = "#059669"
    COLOR_SLOT_HOVER = "#10B981"

# ============================================================================================
#                                    FIELD DEFINITION REGISTRY
# ============================================================================================
class FieldRegistry:
    """Central registry for all extractable fields"""
    
    EFAKTUR_FIELDS = {
        # Header-level fields (will be merged)
        "STATUS": {"label": "STATUS", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NOMOR_FAKTUR": {"label": "NOMOR FAKTUR", "scope": "HEADER", "type": "TEXT", "merge": True},
        "REFERENSI": {"label": "REFERENSI", "scope": "HEADER", "type": "TEXT", "merge": True},
        "KOTA": {"label": "KOTA", "scope": "HEADER", "type": "TEXT", "merge": True},
        "TANGGAL": {"label": "TANGGAL", "scope": "HEADER", "type": "DATE", "merge": True},
        "PENANDATANGAN": {"label": "PENANDATANGAN", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NAMA_PENJUAL": {"label": "NAMA PENJUAL", "scope": "HEADER", "type": "TEXT", "merge": True},
        "ALAMAT_PENJUAL": {"label": "ALAMAT PENJUAL", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NPWP_PENJUAL": {"label": "NPWP PENJUAL", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NAMA_PEMBELI": {"label": "NAMA PEMBELI", "scope": "HEADER", "type": "TEXT", "merge": True},
        "ALAMAT_PEMBELI": {"label": "ALAMAT PEMBELI", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NPWP_PEMBELI": {"label": "NPWP PEMBELI", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NIK": {"label": "NIK", "scope": "HEADER", "type": "TEXT", "merge": True},
        "NOMOR_PASPOR": {"label": "NOMOR PASPOR", "scope": "HEADER", "type": "TEXT", "merge": True},
        "IDENTITAS_LAIN": {"label": "IDENTITAS LAIN", "scope": "HEADER", "type": "TEXT", "merge": True},
        "EMAIL": {"label": "EMAIL", "scope": "HEADER", "type": "TEXT", "merge": True},
        "TOTAL_HARGA": {"label": "TOTAL HARGA", "scope": "HEADER", "type": "NUMBER", "merge": True},
        "TOTAL_POTONGAN": {"label": "TOTAL POTONGAN", "scope": "HEADER", "type": "NUMBER", "merge": True},
        "DPP": {"label": "DPP", "scope": "HEADER", "type": "NUMBER", "merge": True},
        "PPN": {"label": "PPN", "scope": "HEADER", "type": "NUMBER", "merge": True},
        "TOTAL_PPNBM": {"label": "TOTAL PPNBM", "scope": "HEADER", "type": "NUMBER", "merge": True},
        
        # Detail-level fields (per item)
        "NO": {"label": "NO", "scope": "DETAIL", "type": "NUMBER", "merge": False},
        "KODE_BARANG": {"label": "KODE BARANG", "scope": "DETAIL", "type": "TEXT", "merge": False},
        "NAMA_BARANG": {"label": "NAMA BARANG", "scope": "DETAIL", "type": "TEXT", "merge": False},
        "HARGA": {"label": "HARGA", "scope": "DETAIL", "type": "NUMBER", "merge": False},
        "QTY": {"label": "QTY", "scope": "DETAIL", "type": "NUMBER", "merge": False},
        "SATUAN": {"label": "SATUAN", "scope": "DETAIL", "type": "TEXT", "merge": False},
        "POTONGAN": {"label": "POTONGAN", "scope": "DETAIL", "type": "NUMBER", "merge": False},
        "TARIF_PPNBM": {"label": "TARIF PPNBM", "scope": "DETAIL", "type": "TEXT", "merge": False},
        "PPNBM": {"label": "PPNBM", "scope": "DETAIL", "type": "NUMBER", "merge": False},
        "TOTAL_HARGA_BARANG": {"label": "TOTAL HARGA BARANG", "scope": "DETAIL", "type": "NUMBER", "merge": False}
    }
    
    # Default mapping A-AD (30 columns)
    DEFAULT_MAPPING = {
        "A": "NO", "B": "STATUS", "C": "NOMOR_FAKTUR", "D": "REFERENSI",
        "E": "KOTA", "F": "TANGGAL", "G": "PENANDATANGAN", "H": "NAMA_PENJUAL",
        "I": "ALAMAT_PENJUAL", "J": "NPWP_PENJUAL", "K": "NAMA_PEMBELI",
        "L": "ALAMAT_PEMBELI", "M": "NPWP_PEMBELI", "N": "NIK",
        "O": "NOMOR_PASPOR", "P": "IDENTITAS_LAIN", "Q": "EMAIL",
        "R": "KODE_BARANG", "S": "NAMA_BARANG", "T": "HARGA",
        "U": "QTY", "V": "SATUAN", "W": "POTONGAN",
        "X": "TARIF_PPNBM", "Y": "PPNBM", "Z": "TOTAL_HARGA_BARANG",
        "AA": "TOTAL_HARGA", "AB": "TOTAL_POTONGAN", "AC": "DPP", "AD": "PPN",
    }
    
    @classmethod
    def get_all_column_letters(cls):
        """Get all column letters A-AD"""
        single = list(string.ascii_uppercase)
        double = ['A' + c for c in string.ascii_uppercase[:4]]
        return single + double
    
    @classmethod
    def get_available_fields(cls, current_mapping):
        """Get fields that are not yet mapped"""
        mapped_fields = set(current_mapping.values())
        mapped_fields.discard(None)
        
        all_fields = set(cls.EFAKTUR_FIELDS.keys())
        return sorted(all_fields - mapped_fields)

# ============================================================================================
#                                    STATUS DETECTION & EXTRACTORS
# ============================================================================================
def detect_efaktur_status(pdf_path):
    """Detect e-Faktur status by hashing first image"""
    HASH_STATUS_MAP = {
        "c31e32c63de10ee6": "amended",
        "969632c6595b4e6c": "canceled",
        "8000000000000000": "credited",
    }
    
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        images = page.get_images(full=True)
        if not images:
            return "unknown"
        
        xref = images[0][0]
        pix = fitz.Pixmap(doc, xref)
        if pix.n > 4:
            pix = fitz.Pixmap(fitz.csRGB, pix)
        
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        h = str(imagehash.phash(img))
        
        status = HASH_STATUS_MAP.get(h, "unknown")
        if h == "8000000000000000":
            return "approved"
        
        doc.close()
        return status
    except Exception as e:
        print(f"⚠️ Error detecting status: {e}")
        return "unknown"

class EFakturExtractor:
    """Handles e-Faktur PDF extraction with field mapping support"""
    
    BULAN_MAP = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12
    }
    
    @staticmethod
    def parse_tanggal_teks(tanggal_str):
        """Convert '20 Desember 2025' to datetime object"""
        if not tanggal_str or tanggal_str == "":
            return None
        try:
            parts = str(tanggal_str).lower().split()
            if len(parts) != 3:
                return None
            hari = int(parts[0])
            bulan = EFakturExtractor.BULAN_MAP.get(parts[1])
            tahun = int(parts[2])
            if not bulan:
                return None
            return datetime(tahun, bulan, hari)
        except:
            return None
    
    @staticmethod
    def clean_number(value):
        """Convert Indonesian formatted number to integer"""
        if not value:
            return 0
        s = str(value).strip().replace("Rp", "").strip()
        if "," in s:
            s = s.split(",", 1)[0]
        s = s.replace(".", "")
        try:
            return int(s)
        except:
            return 0
    
    @staticmethod
    def find_line_start(lines, keyword):
        for i, l in enumerate(lines):
            if l.startswith(keyword):
                return i, l
        return None, None
    
    @staticmethod
    def find_contains(lines, keyword):
        for i, l in enumerate(lines):
            if keyword in l:
                return i, l
        return None, None
    
    @classmethod
    def extract_all_data(cls, pdf_path, faktur_counter=1, detect_status=True):
        """Extract ALL possible data from PDF"""
        status = "unknown"
        if detect_status:
            status = detect_efaktur_status(pdf_path)
        
        reader = PyPDF2.PdfReader(pdf_path)
        text = ""
        for p in reader.pages:
            text += p.extract_text() + "\n"
        
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        all_data = {"STATUS": status}
        
        # Extract all fields (same logic as before)
        _, line = cls.find_contains(lines, "Kode dan Nomor Seri Faktur Pajak:")
        all_data["NOMOR_FAKTUR"] = re.findall(r"\d+", line)[-1] if line else ""
        
        _, line = cls.find_contains(lines, "(Referensi:")
        all_data["REFERENSI"] = re.findall(r"\(Referensi:\s*([A-Z0-9]+)", line)[0] if line else ""
        
        idx, _ = cls.find_contains(lines, "secara elektronik sehingga tidak diperlukan tanda tangan basah pada Faktur Pajak ini.")
        kota = tanggal = ""
        if idx is not None:
            kota_line = lines[idx + 1]
            kota = kota_line.split(",")[0].strip()
            tanggal_raw = " ".join(lines[idx + 1:idx + 3])
            tgl = re.search(r"(\d{1,2}\s+[A-Za-z]+\s+\d{4})", tanggal_raw)
            tanggal = tgl.group(1) if tgl else ""
        
        all_data["KOTA"] = kota
        all_data["TANGGAL"] = tanggal
        
        idx, _ = cls.find_contains(lines, "Ditandatangani secara elektronik")
        all_data["PENANDATANGAN"] = lines[idx + 1] if idx else ""
        
        idx, _ = cls.find_contains(lines, "Pengusaha Kena Pajak:")
        all_data["NAMA_PENJUAL"] = lines[idx + 1].replace("Nama :", "").strip() if idx else ""
        
        alamat_penjual = []
        if idx:
            for i in range(idx + 2, len(lines)):
                if lines[i].startswith("NPWP"):
                    break
                alamat_penjual.append(lines[i].replace("Alamat :", "").strip())
        all_data["ALAMAT_PENJUAL"] = " ".join(alamat_penjual)
        
        npwps = [l.replace("NPWP :", "").strip() for l in lines if l.startswith("NPWP")]
        all_data["NPWP_PENJUAL"] = npwps[0] if len(npwps) > 0 else ""
        all_data["NPWP_PEMBELI"] = npwps[1] if len(npwps) > 1 else ""
        
        idx, _ = cls.find_contains(lines, "Pembeli Barang Kena Pajak")
        all_data["NAMA_PEMBELI"] = lines[idx + 1].replace("Nama :", "").strip() if idx else ""
        
        alamat_pembeli = []
        if idx:
            for i in range(idx + 2, len(lines)):
                if lines[i].startswith("NPWP"):
                    break
                alamat_pembeli.append(lines[i].replace("Alamat :", "").strip())
        alamat = " ".join(alamat_pembeli)
        alamat = re.sub(r"#\d+$", "", alamat).strip()
        all_data["ALAMAT_PEMBELI"] = alamat
        
        for k in ["NIK", "Nomor Paspor", "Identitas Lain", "Email"]:
            _, l = cls.find_line_start(lines, k)
            field_key = k.upper().replace(" ", "_")
            all_data[field_key] = l.split(":", 1)[1].strip() if l else ""
        
        for key, prefix in {
            "TOTAL_HARGA": "Harga Jual / Penggantian / Uang Muka / Termin",
            "TOTAL_POTONGAN": "Dikurangi Potongan Harga",
            "DPP": "Dasar Pengenaan Pajak",
            "PPN": "Jumlah PPN (Pajak Pertambahan Nilai) ",
            "TOTAL_PPNBM": "Jumlah PPnBM (Pajak Penjualan atas Barang Mewah) "
        }.items():
            _, l = cls.find_line_start(lines, prefix)
            all_data[key] = l.replace(prefix, "").strip() if l else ""
        
        # Extract detail items
        detail_items = []
        i = 0
        while i < len(lines):
            if re.match(r"\d+\s+\d{6}", lines[i]):
                kode = re.findall(r"\d{6}", lines[i])[0]
                nama = re.sub(r"^\d+\s+\d{6}", "", lines[i]).strip()
                
                harga_line = lines[i + 1]
                harga = harga_line.split()[1]
                qty = harga_line.split()[3]
                satuan = harga_line.split()[4]
                potongan = lines[i + 2].split()[-1]
                
                ppnbm_line = lines[i + 3]
                tarif = re.search(r"\((.*?)\)", ppnbm_line).group(1)
                rp = re.findall(r"Rp\s*([\d.,]+)", ppnbm_line)[0]
                
                if "," in rp:
                    after_comma = rp.split(",", 1)[1]
                    total = after_comma[2:] if len(after_comma) > 2 else ""
                else:
                    total = rp
                
                detail_items.append({
                    "KODE_BARANG": kode, "NAMA_BARANG": nama, "HARGA": harga,
                    "QTY": qty, "SATUAN": satuan, "POTONGAN": potongan,
                    "TARIF_PPNBM": tarif, "PPNBM": rp[:4], "TOTAL_HARGA_BARANG": total
                })
                i += 4
            else:
                i += 1
        
        return all_data, detail_items, faktur_counter
    
    @classmethod
    def extract(cls, pdf_path, faktur_counter=1, detect_status=True, field_mapping=None):
        """Extract e-Faktur data based on field mapping"""
        if field_mapping is None:
            field_mapping = FieldRegistry.DEFAULT_MAPPING
        
        header_data, detail_items, counter = cls.extract_all_data(pdf_path, faktur_counter, detect_status)
        active_fields = set(field_mapping.values())
        active_fields.discard(None)
        
        rows = []
        for detail_item in detail_items:
            row = {}
            if "NO" in active_fields:
                row["NO"] = counter
            
            for field_id in active_fields:
                field_info = FieldRegistry.EFAKTUR_FIELDS.get(field_id)
                if field_info and field_info["scope"] == "HEADER":
                    row[field_id] = header_data.get(field_id, "")
            
            for field_id in active_fields:
                field_info = FieldRegistry.EFAKTUR_FIELDS.get(field_id)
                if field_info and field_info["scope"] == "DETAIL":
                    row[field_id] = detail_item.get(field_id, "")
            
            rows.append(row)
        
        ordered_columns = []
        for col in FieldRegistry.get_all_column_letters():
            field_id = field_mapping.get(col)
            if field_id and field_id in active_fields:
                ordered_columns.append(field_id)
        
        df = pd.DataFrame(rows, columns=ordered_columns)
        return df

class BUPOTExtractor:
    """Handles BUPOT PDF extraction"""
    
    BULAN_MAP = {
        "januari": 1, "februari": 2, "maret": 3, "april": 4,
        "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
        "september": 9, "oktober": 10, "november": 11, "desember": 12
    }
    
    @staticmethod
    def ubah_tanggal(teks_tanggal):
        """Convert date text to datetime"""
        if not teks_tanggal:
            return ""
        teks_tanggal = teks_tanggal.strip().lower()
        match = re.match(r"(\d{1,2})\s+([a-z]+)\s+(\d{4})", teks_tanggal)
        if match:
            hari = int(match.group(1))
            bulan = BUPOTExtractor.BULAN_MAP.get(match.group(2))
            tahun = int(match.group(3))
            if bulan:
                return datetime(tahun, bulan, hari)
        match = re.match(r"(\d{2})-(\d{4})", teks_tanggal)
        if match:
            bulan = int(match.group(1))
            tahun = int(match.group(2))
            return datetime(tahun, bulan, 1)
        return ""
    
    @staticmethod
    def bersihkan_angka(teks_angka):
        """Clean number text"""
        if not teks_angka:
            return ""
        angka = re.sub(r"[^\d]", "", teks_angka)
        return int(angka) if angka else ""
    
    @staticmethod
    def deteksi_jenis_dokumen(teks):
        """Detect document type (BPPU or BP21)"""
        lines = teks.splitlines()
        for i, baris in enumerate(lines):
            if "UNIFIKASI BERFORMAT STANDAR" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip().replace(" ", "")
                    if next_line == "BPPU":
                        return "BPPU"
            if "PASAL 21 YANG BERSIFAT FINAL" in baris:
                kata_terakhir = baris.strip().split()[-1]
                if kata_terakhir == "BP21":
                    return "BP21"
        return None
    
    @classmethod
    def ekstrak_bppu(cls, teks):
        """Extract BPPU data"""
        lines = teks.splitlines()
        
        nomor_bukti = masa_pajak = sifat_bukti = status_bukti = ""
        for i, baris in enumerate(lines):
            if "A. IDENTITAS WAJIB PAJAK YANG DIPOTONG DAN/ATAU DIPUNGUT PPh ATAU PENERIMA PENGHASILAN" in baris:
                if i > 0:
                    prev_line = lines[i - 1].strip()
                    parts = prev_line.split()
                    if len(parts) >= 4:
                        nomor_bukti = parts[0]
                        masa_pajak = parts[1]
                        sifat_bukti = " ".join(parts[2:-1])
                        status_bukti = parts[-1]
                break
        
        masa_pajak_obj = cls.ubah_tanggal(masa_pajak)
        
        kode_objek = ""
        for i, baris in enumerate(lines):
            if "B.3 B.4 B.5 B.6 B.7" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    kode_objek = next_line.split()[0]
                break
        
        dpp = pph = ""
        for i, baris in enumerate(lines):
            if "B.8 Dokumen Dasar Bukti" in baris:
                if i >= 2:
                    target_line = lines[i - 2].strip()
                    angka_semua = re.findall(r'\d[\d\.]*', target_line)
                    if len(angka_semua) >= 3:
                        dpp = angka_semua[-3]
                        pph = angka_semua[-1]
                break
        
        npwp_pemotong = ""
        for baris in lines:
            if "C.1 NPWP / NIK :" in baris:
                npwp_pemotong = baris.split(":")[-1].strip()
                break
        
        nama_pemotong = ""
        for i, baris in enumerate(lines):
            if "C.3 NAMA PEMOTONG DAN/ATAU PEMUNGUT" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    if "PPh:" in next_line:
                        nama_pemotong = next_line.split(":")[-1].strip()
                break
        
        tanggal_bukti = ""
        for baris in lines:
            if "C.4 TANGGAL :" in baris:
                tanggal_bukti = baris.split(":")[-1].strip()
                break
        
        tanggal_bukti_obj = cls.ubah_tanggal(tanggal_bukti)
        
        return {
            'JENIS DOKUMEN': 'BPPU',
            'NOMOR BUKTI POTONG': nomor_bukti,
            'MASA PAJAK': masa_pajak_obj,
            'SIFAT BUKTI POTONG': sifat_bukti,
            'STATUS BUKTI POTONG': status_bukti,
            'KODE OBJEK PAJAK': kode_objek,
            'DPP': cls.bersihkan_angka(dpp),
            'PPH YANG DIPOTONG': cls.bersihkan_angka(pph),
            'NPWP PEMOTONG': npwp_pemotong,
            'NAMA PEMOTONG': nama_pemotong,
            'TANGGAL BUKTI POTONG': tanggal_bukti_obj
        }
    
    @classmethod
    def ekstrak_bp21(cls, teks):
        """Extract BP21 data"""
        lines = teks.splitlines()
        
        nomor_bukti = masa_pajak = sifat_bukti = status_bukti = ""
        for i, baris in enumerate(lines):
            if "NOMOR BUKTI PEMOTONGAN MASA PAJAK SIFAT PEMOTONGAN STATUS BUKTI PEMOTONGAN" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    parts = next_line.split()
                    if len(parts) >= 4:
                        nomor_bukti = parts[0]
                        masa_pajak = parts[1]
                        sifat_bukti = " ".join(parts[2:-1])
                        status_bukti = parts[-1]
                break
        
        masa_pajak_obj = cls.ubah_tanggal(masa_pajak)
        
        kode_objek = ""
        for i, baris in enumerate(lines):
            if "B.2 B.3 B.4 B.5 B.6 B.7" in baris:
                if i + 1 < len(lines):
                    next_line = lines[i + 1].strip()
                    kode_objek = next_line.split()[0]
                break
        
        dpp = pph = ""
        for i, baris in enumerate(lines):
            if "B.8 Dokumen Referensi Jenis Dokumen" in baris or "B.8 Dokumen Referensi" in baris:
                if i >= 1:
                    target_line = lines[i - 1].strip()
                    angka_bersih = re.sub(r'[A-Za-z\s]+', ' ', target_line)
                    angka_semua = angka_bersih.split()
                    if len(angka_semua) >= 4:
                        dpp = angka_semua[0]
                        pph = angka_semua[-1]
                break
        
        npwp_pemotong = ""
        for baris in lines:
            if "C.1 NPWP/NIK :" in baris:
                npwp_pemotong = baris.split(":")[-1].strip()
                break
        
        nama_pemotong = ""
        for baris in lines:
            if "C.3 Nama Pemotong :" in baris:
                nama_pemotong = baris.split(":")[-1].strip()
                break
        
        tanggal_bukti = ""
        for baris in lines:
            if "C.4 Tanggal :" in baris:
                tanggal_bukti = baris.split(":")[-1].strip()
                break
        
        tanggal_bukti_obj = cls.ubah_tanggal(tanggal_bukti)
        
        return {
            'JENIS DOKUMEN': 'BP21',
            'NOMOR BUKTI POTONG': nomor_bukti,
            'MASA PAJAK': masa_pajak_obj,
            'SIFAT BUKTI POTONG': sifat_bukti,
            'STATUS BUKTI POTONG': status_bukti,
            'KODE OBJEK PAJAK': kode_objek,
            'DPP': cls.bersihkan_angka(dpp),
            'PPH YANG DIPOTONG': cls.bersihkan_angka(pph),
            'NPWP PEMOTONG': npwp_pemotong,
            'NAMA PEMOTONG': nama_pemotong,
            'TANGGAL BUKTI POTONG': tanggal_bukti_obj
        }
    
    @classmethod
    def extract(cls, pdf_path):
        """Extract BUPOT data from PDF"""
        reader = PyPDF2.PdfReader(pdf_path)
        pdf_text = ""
        for page in reader.pages:
            pdf_text += page.extract_text() or ""
        
        jenis_dokumen = cls.deteksi_jenis_dokumen(pdf_text)
        
        if jenis_dokumen == "BPPU":
            data = cls.ekstrak_bppu(pdf_text)
        elif jenis_dokumen == "BP21":
            data = cls.ekstrak_bp21(pdf_text)
        else:
            raise ValueError("Jenis dokumen tidak dikenali")
        
        return pd.DataFrame([data])

# ============================================================================================
#                                    FIELD MAPPING CONTROLLER
# ============================================================================================
class FieldMappingController:
    """Central controller for field mapping state and logic"""
    
    def __init__(self, main_window):
        self.main_window = main_window
        
        # STATE (Single Source of Truth)
        self.ordered_columns = FieldRegistry.get_all_column_letters()
        self.field_mapping = {}
        self.available_fields = []
        self.current_dragged_field = None
        
        # Widget tracking for removal
        self.field_widgets_in_pool = {}
        self.field_widgets_in_slots = {}
        
        self._load_default_mapping()
        self._recalculate_available_fields()
    
    def _load_default_mapping(self):
        for col, field_id in FieldRegistry.DEFAULT_MAPPING.items():
            self.field_mapping[col] = field_id
    
    def _recalculate_available_fields(self):
        mapped = set(self.field_mapping.values())
        mapped.discard(None)
        all_fields = set(FieldRegistry.EFAKTUR_FIELDS.keys())
        self.available_fields = sorted(all_fields - mapped)
    
    def start_drag(self, field_id, source_widget):
        """Start drag from pool or slot"""
        self.current_dragged_field = field_id
        self.drag_source_widget = source_widget
    
    def drop_to_slot(self, column_letter):
        """Drop field to slot"""
        if not self.current_dragged_field:
            return False
        
        # Check if slot occupied
        if self.field_mapping.get(column_letter):
            self.current_dragged_field = None
            return False
        
        # Map field
        self.field_mapping[column_letter] = self.current_dragged_field
        
        # Clean up source widget if from pool
        if hasattr(self, 'drag_source_widget'):
            if self.drag_source_widget.parent() == self.main_window.field_pool_widget:
                self.drag_source_widget.setParent(None)
                self.drag_source_widget.deleteLater()
        
        self.current_dragged_field = None
        self._recalculate_available_fields()
        self.main_window.refresh_mapping_ui()
        return True
    
    def remove_from_slot(self, column_letter):
        """Remove field and REFLOW (shift left)"""
        if not self.field_mapping.get(column_letter):
            return
        
        try:
            idx = self.ordered_columns.index(column_letter)
        except ValueError:
            return
        
        # Store removed field
        removed_field = self.field_mapping[column_letter]
        
        # Clear slot
        self.field_mapping[column_letter] = None
        
        # REFLOW: Shift all columns to the right → left
        for i in range(idx, len(self.ordered_columns) - 1):
            src_col = self.ordered_columns[i + 1]
            dst_col = self.ordered_columns[i]
            self.field_mapping[dst_col] = self.field_mapping.get(src_col)
        
        last_col = self.ordered_columns[-1]
        self.field_mapping[last_col] = None
        
        self._recalculate_available_fields()
        self.main_window.refresh_mapping_ui()
    
    def reset_to_default(self):
        self.field_mapping = {}
        self._load_default_mapping()
        self.current_dragged_field = None
        self._recalculate_available_fields()
        self.main_window.refresh_mapping_ui()
    
    def get_mapping_snapshot(self):
        return self.field_mapping.copy()
    
    def load_from_config(self, config_mapping):
        if config_mapping:
            self.field_mapping = config_mapping.copy()
            self._recalculate_available_fields()

# ============================================================================================
#                                    UI COMPONENTS
# ============================================================================================
class DraggableField(QFrame):
    """Draggable field widget with TRUE Qt drag & drop + PIXMAP"""
    
    def __init__(self, field_id, label, controller):
        super().__init__()
        self.field_id = field_id
        self.controller = controller
        self.is_selected = False
        
        self.setFixedHeight(40)
        self.setCursor(QCursor(Qt.OpenHandCursor))
        self.setup_ui(label)
        self.set_style(False)
    
    def setup_ui(self, label):
        layout = QVBoxLayout()
        layout.setContentsMargins(8, 6, 8, 6)
        
        self.label = QLabel(label)
        self.label.setFont(QFont("Arial", 10, QFont.Bold))
        self.label.setAlignment(Qt.AlignCenter)
        self.label.setStyleSheet("color: white;")
        
        layout.addWidget(self.label)
        self.setLayout(layout)
    
    def set_style(self, selected):
        self.is_selected = selected
        if selected:
            self.setStyleSheet(f"""
                QFrame {{
                    background-color: {Config.COLOR_FIELD_SELECTED};
                    border: 2px solid #F59E0B;
                    border-radius: 8px;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                QFrame {{
                    background-color: {Config.COLOR_FIELD};
                    border: 1px solid #3B82F6;
                    border-radius: 8px;
                }}
                QFrame:hover {{
                    background-color: {Config.COLOR_FIELD_HOVER};
                }}
            """)
    
    def mousePressEvent(self, event):
        """Start drag operation with PIXMAP animation"""
        if event.button() == Qt.LeftButton:
            # Visual feedback
            self.setCursor(QCursor(Qt.ClosedHandCursor))
            
            # Notify controller
            self.controller.start_drag(self.field_id, self)
            
            # Start Qt drag operation
            drag = QDrag(self)
            mime_data = QMimeData()
            mime_data.setText(self.field_id)
            drag.setMimeData(mime_data)
            
            # 🔥 ADD PIXMAP for visual drag feedback
            pixmap = self.grab()
            drag.setPixmap(pixmap)
            drag.setHotSpot(pixmap.rect().center())
            
            # Execute drag
            drag.exec(Qt.MoveAction)
            
            self.setCursor(QCursor(Qt.OpenHandCursor))

class DropSlot(QFrame):
    """Drop slot that can CONTAIN DraggableField widgets"""
    
    MIN_WIDTH = 70
    MAX_WIDTH = 250
    CHAR_WIDTH = 7
    PADDING = 20
    
    def __init__(self, column_letter, controller):
        super().__init__()
        self.column_letter = column_letter
        self.controller = controller
        self.current_width = self.MIN_WIDTH
        self.contained_widget = None  # 🔥 Widget-based architecture
        
        self.setFixedSize(self.MIN_WIDTH, 60)
        self.setCursor(QCursor(Qt.PointingHandCursor))
        
        # Enable drop events
        self.setAcceptDrops(True)
        
        self.setup_ui()
        self.display_empty()
    
    def setup_ui(self):
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(5, 5, 5, 5)
        self.setLayout(self.layout)
    
    def dragEnterEvent(self, event):
        """Accept drag if it has field_id data"""
        if event.mimeData().hasText():
            event.acceptProposedAction()
            # Visual feedback
            self.setStyleSheet(f"""
                QFrame {{
                    background-color: {Config.COLOR_SLOT_HOVER};
                    border: 3px solid #34D399;
                    border-radius: 8px;
                }}
            """)
    
    def dragLeaveEvent(self, event):
        """Remove visual feedback when drag leaves"""
        # Restore original style
        if self.contained_widget:
            self.setStyleSheet(f"""
                QFrame {{
                    background-color: {Config.COLOR_SLOT_FILLED};
                    border: 2px solid #10B981;
                    border-radius: 8px;
                }}
                QFrame:hover {{
                    border-color: #34D399;
                }}
            """)
        else:
            self.setStyleSheet(f"""
                QFrame {{
                    background-color: {Config.COLOR_SLOT_EMPTY};
                    border: 2px solid #4B5563;
                    border-radius: 8px;
                }}
                QFrame:hover {{
                    border-color: #6B7280;
                }}
            """)
    
    def dropEvent(self, event):
        """Handle drop - execute placement"""
        field_id = event.mimeData().text()
        
        success = self.controller.drop_to_slot(self.column_letter)
        
        if not success and self.controller.field_mapping.get(self.column_letter):
            QMessageBox.warning(
                self,
                "Slot Terisi",
                f"Kolom {self.column_letter} sudah terisi!\n\nKlik kanan untuk menghapus."
            )
        
        event.acceptProposedAction()
    
    def display_field_widget(self, field_id, field_label):
        """Display field as contained widget"""
        # Clear any existing widget
        if self.contained_widget:
            self.contained_widget.setParent(None)
            self.contained_widget.deleteLater()
        
        # Create new draggable field widget
        widget = DraggableField(field_id, field_label, self.controller)
        widget.setFixedHeight(50)
        self.layout.addWidget(widget)
        self.contained_widget = widget
        
        # Calculate width
        new_width = self._calculate_width(field_label)
        self.current_width = new_width
        self.setFixedWidth(new_width)
        
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {Config.COLOR_SLOT_FILLED};
                border: 2px solid #10B981;
                border-radius: 8px;
            }}
            QFrame:hover {{
                border-color: #34D399;
            }}
        """)
        
        return new_width
    
    def display_empty(self):
        """Display empty slot"""
        if self.contained_widget:
            self.contained_widget.setParent(None)
            self.contained_widget.deleteLater()
            self.contained_widget = None
        
        self.current_width = self.MIN_WIDTH
        self.setFixedWidth(self.MIN_WIDTH)
        
        self.setStyleSheet(f"""
            QFrame {{
                background-color: {Config.COLOR_SLOT_EMPTY};
                border: 2px solid #4B5563;
                border-radius: 8px;
            }}
            QFrame:hover {{
                border-color: #6B7280;
            }}
        """)
        
        return self.MIN_WIDTH
    
    def _calculate_width(self, text):
        estimated = len(text) * self.CHAR_WIDTH + self.PADDING
        return max(self.MIN_WIDTH, min(estimated, self.MAX_WIDTH))
    
    def get_width(self):
        return self.current_width
    
    def mousePressEvent(self, event):
        """Right-click to remove (keeps existing functionality)"""
        if event.button() == Qt.RightButton:
            self.controller.remove_from_slot(self.column_letter)

class ColumnHeader(QFrame):
    """Column header that syncs with slot"""
    
    def __init__(self, column_letter):
        super().__init__()
        self.column_letter = column_letter
        self.current_width = 70
        
        self.setFixedSize(70, 30)
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)
        
        label = QLabel(self.column_letter)
        label.setFont(QFont("Arial", 10, QFont.Bold))
        label.setAlignment(Qt.AlignCenter)
        label.setStyleSheet("color: white;")
        
        layout.addWidget(label)
        self.setLayout(layout)
        
        self.setStyleSheet("""
            QFrame {
                background-color: #374151;
                border-radius: 5px;
            }
        """)
    
    def sync_width(self, width):
        self.current_width = width
        self.setFixedWidth(width)

# ============================================================================================
#                                    MAIN WINDOW
# ============================================================================================
class SmartPDFExtractor(QMainWindow):
    
    def __init__(self):
        super().__init__()
        
        # Expiry check
        if datetime.now() > datetime(2026, 5, 2, 23, 59, 59):
            sys.exit()
        
        self.setWindowTitle(Config.WINDOW_TITLE)
        self.setGeometry(100, 100, Config.WINDOW_WIDTH, Config.WINDOW_HEIGHT)
        self.setMinimumSize(Config.MIN_WIDTH, Config.MIN_HEIGHT)
        
        # Variables
        self.selected_files = []
        self.final_df_global = None
        
        # Config
        self.config_file = self._get_config_path()
        self.load_config()
        
        # Controller
        self.mapping_controller = FieldMappingController(self)
        self._apply_loaded_config()
        
        # UI References
        self.drop_slots = {}
        self.column_headers = {}
        self.field_widgets = []
        
        # Setup UI
        self.setup_ui()
        
        # Apply dark theme
        self.set_dark_theme()
    
    def setup_ui(self):
        """Setup main UI"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        main_layout = QHBoxLayout()
        main_layout.setSpacing(0)
        main_layout.setContentsMargins(0, 0, 0, 0)
        
        # Sidebar
        sidebar = self.create_sidebar()
        main_layout.addWidget(sidebar)
        
        # Pages container
        self.pages_widget = QWidget()
        self.pages_layout = QVBoxLayout()
        self.pages_layout.setContentsMargins(0, 0, 0, 0)
        self.pages_widget.setLayout(self.pages_layout)
        
        main_layout.addWidget(self.pages_widget, 1)
        
        central_widget.setLayout(main_layout)
        
        # Create pages
        self.home_page = self.build_home_page()
        self.mapping_page = self.build_field_mapping_page()
        self.efaktur_keluaran_page = self.build_efaktur_page("keluaran")
        self.efaktur_masukan_page = self.build_efaktur_page("masukan")
        self.bupot_page = self.build_bupot_page()
        
        self.pages = [
            self.home_page,
            self.mapping_page,
            self.efaktur_keluaran_page,
            self.efaktur_masukan_page,
            self.bupot_page
        ]
        
        # Add pages to layout
        for page in self.pages:
            self.pages_layout.addWidget(page)
            page.hide()
        
        # Show home page
        self.show_page(0)
    
    def create_sidebar(self):
        """Create sidebar with navigation buttons"""
        sidebar = QFrame()
        sidebar.setFixedWidth(80)
        sidebar.setStyleSheet(f"background-color: {Config.COLOR_SIDEBAR};")
        
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(10, 20, 10, 10)
        
        # Home button
        self.btn_home = QPushButton("🏠")
        self.btn_home.setFixedSize(60, 60)
        self.btn_home.setFont(QFont("Arial", 30))
        self.btn_home.clicked.connect(lambda: self.show_page(0))
        self.style_sidebar_button(self.btn_home)
        layout.addWidget(self.btn_home)
        
        # Mapping button
        self.btn_mapping = QPushButton("🎯")
        self.btn_mapping.setFixedSize(60, 60)
        self.btn_mapping.setFont(QFont("Arial", 30))
        self.btn_mapping.clicked.connect(lambda: self.show_page(1))
        self.style_sidebar_button(self.btn_mapping)
        layout.addWidget(self.btn_mapping)
        
        # E-Faktur Keluaran button
        self.btn_efaktur_keluaran = QPushButton("📤")
        self.btn_efaktur_keluaran.setFixedSize(60, 60)
        self.btn_efaktur_keluaran.setFont(QFont("Arial", 30))
        self.btn_efaktur_keluaran.clicked.connect(lambda: self.show_page(2))
        self.style_sidebar_button(self.btn_efaktur_keluaran)
        layout.addWidget(self.btn_efaktur_keluaran)
        
        # E-Faktur Masukan button
        self.btn_efaktur_masukan = QPushButton("📥")
        self.btn_efaktur_masukan.setFixedSize(60, 60)
        self.btn_efaktur_masukan.setFont(QFont("Arial", 30))
        self.btn_efaktur_masukan.clicked.connect(lambda: self.show_page(3))
        self.style_sidebar_button(self.btn_efaktur_masukan)
        layout.addWidget(self.btn_efaktur_masukan)
        
        # BUPOT button
        self.btn_bupot = QPushButton("📃")
        self.btn_bupot.setFixedSize(60, 60)
        self.btn_bupot.setFont(QFont("Arial", 30))
        self.btn_bupot.clicked.connect(lambda: self.show_page(4))
        self.style_sidebar_button(self.btn_bupot)
        layout.addWidget(self.btn_bupot)
        
        layout.addStretch()
        
        sidebar.setLayout(layout)
        return sidebar
    
    def style_sidebar_button(self, button):
        """Style sidebar navigation button"""
        button.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_BUTTON};
                border: none;
                border-radius: 12px;
                color: white;
            }}
            QPushButton:hover {{
                background-color: {Config.COLOR_BUTTON_HOVER};
            }}
        """)
    
    def build_home_page(self):
        """Build home page"""
        page = QScrollArea()
        page.setWidgetResizable(True)
        page.setStyleSheet(f"background-color: {Config.COLOR_BG_MEDIUM};")
        
        content = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 20, 40, 20)
        
        # Title
        title = QLabel("⬢ Smart PDF AutoExtractor")
        title.setFont(QFont("Arial", 28, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(f"""
            background-color: {Config.COLOR_ACCENT};
            color: black;
            padding: 15px;
            border-radius: 20px;
        """)
        layout.addWidget(title)
        
        # Subtitle
        subtitle = QLabel("Ekstraksi Otomatis PDF ke Excel untuk E-Faktur & BUPOT")
        subtitle.setFont(QFont("Arial", 14))
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: white; padding: 10px;")
        layout.addWidget(subtitle)
        
        # Instructions
        instructions_label = QLabel("Cara Penggunaan")
        instructions_label.setFont(QFont("Arial", 18, QFont.Bold))
        instructions_label.setStyleSheet("color: white;")
        layout.addWidget(instructions_label)
        
        instructions = [
            "1. (OPSIONAL) Atur Field Mapping dengan DRAG & DROP field ke slot",
            "2. Pilih jenis ekstraksi dari menu di kiri",
            "3. Upload file PDF (bisa multiple files)",
            "4. Klik 'Proses Sekarang' untuk ekstraksi",
            "5. Download hasil Excel dengan formatting profesional"
        ]
        
        for ins in instructions:
            label = QLabel(ins)
            label.setStyleSheet("color: white; padding: 5px;")
            layout.addWidget(label)
        
        # Features
        features_label = QLabel("Fitur Utama")
        features_label.setFont(QFont("Arial", 18, QFont.Bold))
        features_label.setStyleSheet("color: white; margin-top: 20px;")
        layout.addWidget(features_label)
        
        features = [
            "✅ TRUE Drag & Drop: Seret field dengan animasi visual (Qt native + pixmap)",
            "✅ Drag Keluar dari Slot: Field di slot bisa di-drag keluar!",
            "✅ Auto-Shift REFLOW: Hapus field → yang lain geser otomatis!",
            "✅ Widget-Based: Field = widget yang bisa pindah pool ↔ slot",
            "✅ E-Faktur: Ekstrak lengkap semua data (Keluaran & Masukan)",
            "✅ BUPOT: Auto-detect BPPU atau BP21",
            "✅ Format Excel profesional dengan merged cells & borders",
            "✅ Thread-safe extraction"
        ]
        
        for feat in features:
            label = QLabel(feat)
            label.setStyleSheet("color: white; padding: 5px;")
            layout.addWidget(label)
        
        layout.addStretch()
        
        # Promo
        promo_frame = QFrame()
        promo_frame.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border: 2px solid {Config.COLOR_WARNING};
            border-radius: 20px;
            padding: 15px;
        """)
        promo_layout = QVBoxLayout()
        
        promo_label = QLabel("Sambil nungguin laper yaaa? Coba snack di 🥰✨Kedai Ayam Warisan 81✨🥰")
        promo_label.setAlignment(Qt.AlignCenter)
        promo_label.setStyleSheet("color: white; font-size: 12px;")
        promo_layout.addWidget(promo_label)
        
        btn_layout = QHBoxLayout()
        
        btn_gofood = QPushButton("🛵 GoFood")
        btn_gofood.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_ERROR};
                color: white;
                border: none;
                border-radius: 15px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #DC2626;
            }}
        """)
        btn_gofood.clicked.connect(lambda: webbrowser.open("https://gofood.link/a/Mtv3P3L"))
        btn_layout.addWidget(btn_gofood)
        
        btn_grabfood = QPushButton("🟢 GrabFood")
        btn_grabfood.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_SUCCESS};
                color: white;
                border: none;
                border-radius: 15px;
                padding: 10px 20px;
                font-weight: bold;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        btn_grabfood.clicked.connect(lambda: webbrowser.open("https://r.grab.com/g/6-20260202_212151_1f6a2784162b40d5bd1c465b0e817b13_MEXMPS-6-C6NAVVMAVTDEFE"))
        btn_layout.addWidget(btn_grabfood)
        
        promo_layout.addLayout(btn_layout)
        
        wa_label = QLabel("📞 089671139111 (Klik untuk menyalin)")
        wa_label.setAlignment(Qt.AlignCenter)
        wa_label.setStyleSheet(f"color: {Config.COLOR_ACCENT}; cursor: pointer;")
        wa_label.mousePressEvent = lambda e: self.copy_to_clipboard("089671139111")
        promo_layout.addWidget(wa_label)
        
        promo_frame.setLayout(promo_layout)
        layout.addWidget(promo_frame)
        
        # Credit
        credit = QLabel("BY RAYMOND FO + CLAUDE")
        credit.setStyleSheet("color: gray;")
        credit.setAlignment(Qt.AlignRight)
        layout.addWidget(credit)
        
        content.setLayout(layout)
        page.setWidget(content)
        
        return page
    
    def build_field_mapping_page(self):
        """Build field mapping page with controller architecture"""
        page = QScrollArea()
        page.setWidgetResizable(True)
        page.setStyleSheet(f"background-color: {Config.COLOR_BG_MEDIUM};")
        
        content = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(10)
        layout.setContentsMargins(20, 15, 20, 15)
        
        # Header
        header = QFrame()
        header.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border: 1px solid #1F2937;
            border-radius: 15px;
        """)
        header_layout = QVBoxLayout()
        
        title = QLabel("🎯 Field Mapping untuk E-Faktur")
        title.setFont(QFont("Arial", 20, QFont.Bold))
        title.setStyleSheet("color: white; padding: 10px;")
        header_layout.addWidget(title)
        
        subtitle = QLabel("DRAG field (dengan animasi!) → DROP ke slot | KLIK KANAN untuk hapus (auto-shift!) | Field di slot BISA di-drag keluar!")
        subtitle.setStyleSheet("color: #94A3B8; padding: 5px 10px;")
        header_layout.addWidget(subtitle)
        
        header.setLayout(header_layout)
        layout.addWidget(header)
        
        # Info box
        info = QFrame()
        info.setStyleSheet(f"""
            background-color: {Config.COLOR_BUTTON};
            border: 1px solid #3B82F6;
            border-radius: 10px;
            padding: 10px;
        """)
        info_layout = QVBoxLayout()
        
        tips = [
            "💡 DRAG field dari pool (tahan + seret mouse) → ada animasi!",
            "💡 DROP ke slot kolom (lepas mouse di slot)",
            "💡 DRAG field dari slot → kembali ke pool atau ke slot lain!",
            "💡 KLIK KANAN slot untuk HAPUS → auto REFLOW!",
            "💡 Width auto-sync antara header & slot"
        ]
        
        for tip in tips:
            label = QLabel(tip)
            label.setStyleSheet("color: white; font-size: 11px;")
            info_layout.addWidget(label)
        
        info.setLayout(info_layout)
        layout.addWidget(info)
        
        # Columns and slots container
        mapping_container = QFrame()
        mapping_container.setStyleSheet(f"background-color: {Config.COLOR_BG_DARK}; border-radius: 10px;")
        mapping_layout = QVBoxLayout()
        
        # Scroll area for horizontal columns
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFixedHeight(140)
        
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout()
        scroll_layout.setContentsMargins(10, 10, 10, 10)
        
        # Grid for headers and slots
        grid = QGridLayout()
        grid.setSpacing(2)
        
        columns = FieldRegistry.get_all_column_letters()
        for i, col in enumerate(columns):
            # Header
            header = ColumnHeader(col)
            grid.addWidget(header, 0, i)
            self.column_headers[col] = header
            
            # Slot
            slot = DropSlot(col, self.mapping_controller)
            grid.addWidget(slot, 1, i)
            self.drop_slots[col] = slot
        
        scroll_layout.addLayout(grid)
        scroll_content.setLayout(scroll_layout)
        scroll.setWidget(scroll_content)
        
        mapping_layout.addWidget(scroll)
        mapping_container.setLayout(mapping_layout)
        layout.addWidget(mapping_container)
        
        # Buttons
        btn_layout = QHBoxLayout()
        
        btn_reset = QPushButton("🔄 Reset Default")
        btn_reset.setFixedHeight(40)
        btn_reset.clicked.connect(self.reset_mapping)
        btn_reset.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_ERROR};
                color: white;
                border: none;
                border-radius: 20px;
                font-weight: bold;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                background-color: #DC2626;
            }}
        """)
        btn_layout.addWidget(btn_reset)
        
        btn_save = QPushButton("💾 Save Mapping")
        btn_save.setFixedHeight(40)
        btn_save.clicked.connect(self.save_mapping)
        btn_save.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_SUCCESS};
                color: white;
                border: none;
                border-radius: 20px;
                font-weight: bold;
                padding: 10px 20px;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        btn_layout.addWidget(btn_save)
        
        layout.addLayout(btn_layout)
        
        # Field Pool
        pool_frame = QFrame()
        pool_frame.setStyleSheet(f"background-color: {Config.COLOR_BG_DARK}; border-radius: 10px;")
        pool_layout = QVBoxLayout()
        
        pool_label = QLabel("📦 Field Pool (Drag field untuk mulai):")
        pool_label.setFont(QFont("Arial", 12, QFont.Bold))
        pool_label.setStyleSheet("color: white; padding: 10px;")
        pool_layout.addWidget(pool_label)
        
        # Scrollable field pool
        self.field_pool_scroll = QScrollArea()
        self.field_pool_scroll.setWidgetResizable(True)
        self.field_pool_scroll.setStyleSheet(f"background-color: {Config.COLOR_BG_MEDIUM}; border-radius: 10px;")
        
        self.field_pool_widget = QWidget()
        self.field_pool_layout = QVBoxLayout()
        self.field_pool_layout.setSpacing(5)
        self.field_pool_widget.setLayout(self.field_pool_layout)
        
        self.field_pool_scroll.setWidget(self.field_pool_widget)
        pool_layout.addWidget(self.field_pool_scroll)
        
        pool_frame.setLayout(pool_layout)
        layout.addWidget(pool_frame, 1)
        
        content.setLayout(layout)
        page.setWidget(content)
        
        # Initial UI refresh
        QTimer.singleShot(100, self.refresh_mapping_ui)
        
        return page
    
    def refresh_mapping_ui(self):
        """CENTRAL UI REFRESH - called by controller after state changes"""
        # 1. Refresh slots and headers
        for col, slot in self.drop_slots.items():
            field_id = self.mapping_controller.field_mapping.get(col)
            
            if field_id:
                field_info = FieldRegistry.EFAKTUR_FIELDS[field_id]
                new_width = slot.display_field_widget(field_id, field_info["label"])
                self.column_headers[col].sync_width(new_width)
            else:
                new_width = slot.display_empty()
                self.column_headers[col].sync_width(new_width)
        
        # 2. Rebuild field pool
        # Clear existing widgets
        while self.field_pool_layout.count():
            item = self.field_pool_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        self.field_widgets = []
        
        # Create new widgets from available_fields list
        row_widget = None
        row_layout = None
        items_in_row = 0
        max_items_per_row = 4
        
        for field_id in self.mapping_controller.available_fields:
            field_info = FieldRegistry.EFAKTUR_FIELDS[field_id]
            
            if items_in_row == 0:
                row_widget = QWidget()
                row_layout = QHBoxLayout()
                row_layout.setSpacing(4)
                row_layout.setContentsMargins(0, 0, 0, 0)
                row_widget.setLayout(row_layout)
                self.field_pool_layout.addWidget(row_widget)
            
            field_widget = DraggableField(field_id, field_info["label"], self.mapping_controller)
            row_layout.addWidget(field_widget)
            
            self.field_widgets.append(field_widget)
            
            items_in_row += 1
            if items_in_row >= max_items_per_row:
                items_in_row = 0
        
        # Add stretch to last row if needed
        if row_layout and items_in_row > 0:
            row_layout.addStretch()
        
        self.field_pool_layout.addStretch()
    
    def reset_mapping(self):
        """Reset to default mapping"""
        reply = QMessageBox.question(
            self, "Konfirmasi", "Reset mapping ke default?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply == QMessageBox.Yes:
            self.mapping_controller.reset_to_default()
            QMessageBox.information(self, "Berhasil", "Mapping telah direset ke default!")
    
    def save_mapping(self):
        """Save current mapping"""
        active_fields = [f for f in self.mapping_controller.field_mapping.values() if f]
        if not active_fields:
            QMessageBox.warning(self, "Peringatan", "Minimal satu field harus dipilih!")
            return
        
        self.save_config()
        QMessageBox.information(
            self, "Berhasil",
            f"Mapping tersimpan!\n\n{len(active_fields)} field aktif akan digunakan untuk ekstraksi."
        )
    
    def build_efaktur_page(self, mode):
        """Build E-Faktur extraction page
        
        Args:
            mode: 'keluaran' or 'masukan'
        """
        page = QScrollArea()
        page.setWidgetResizable(True)
        page.setStyleSheet(f"background-color: {Config.COLOR_BG_MEDIUM};")
        
        content = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 20, 40, 20)
        
        # Title
        title_text = f"🚀 E-Faktur {'Keluaran' if mode == 'keluaran' else 'Masukan'} Extractor"
        title = QLabel(title_text)
        title.setFont(QFont("Arial", 24, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(f"""
            background-color: {Config.COLOR_ACCENT};
            color: black;
            padding: 15px;
            border-radius: 20px;
        """)
        layout.addWidget(title)
        
        # Subtitle
        subtitle = QLabel("Upload PDF dan ekstrak data ke Excel (menggunakan Field Mapping yang aktif)")
        subtitle.setFont(QFont("Arial", 12))
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: white; padding: 10px;")
        layout.addWidget(subtitle)
        
        # Upload Frame
        upload_frame = QFrame()
        upload_frame.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border: 2px solid {Config.COLOR_ACCENT};
            border-radius: 20px;
            padding: 20px;
        """)
        upload_layout = QVBoxLayout()
        
        upload_label = QLabel("Upload File PDF")
        upload_label.setStyleSheet("color: white; font-size: 14px;")
        upload_label.setAlignment(Qt.AlignCenter)
        upload_layout.addWidget(upload_label)
        
        # File button
        file_button = QPushButton("📁 PILIH FILE PDF")
        file_button.setFixedHeight(45)
        file_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_ACCENT};
                color: black;
                border: none;
                border-radius: 22px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #67E8F9;
            }}
        """)
        file_button.clicked.connect(lambda: self.select_files(mode))
        upload_layout.addWidget(file_button)
        
        upload_frame.setLayout(upload_layout)
        layout.addWidget(upload_frame)
        
        # File List
        file_frame = QFrame()
        file_frame.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border-radius: 15px;
            padding: 15px;
        """)
        file_layout = QVBoxLayout()
        
        file_list = QTextEdit()
        file_list.setReadOnly(True)
        file_list.setMinimumHeight(150)
        file_list.setStyleSheet("""
            background-color: #0F172A;
            color: white;
            border: 1px solid #374151;
            border-radius: 8px;
            padding: 10px;
        """)
        file_layout.addWidget(file_list)
        
        file_frame.setLayout(file_layout)
        layout.addWidget(file_frame, 1)
        
        # Process Button
        process_btn = QPushButton("⚡ PROSES SEKARANG")
        process_btn.setFixedHeight(50)
        process_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_SUCCESS};
                color: white;
                border: none;
                border-radius: 25px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        process_btn.clicked.connect(lambda: self.process_extraction(mode))
        layout.addWidget(process_btn)
        
        # Status Label
        status = QLabel("Status: Menunggu file...")
        status.setStyleSheet("color: gray; font-size: 12px;")
        status.setAlignment(Qt.AlignCenter)
        layout.addWidget(status)
        
        # Download Button
        download_btn = QPushButton("⬇ DOWNLOAD HASIL EXCEL")
        download_btn.setFixedHeight(50)
        download_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #A78BFA;
                color: white;
                border: none;
                border-radius: 25px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #C4B5FD;
            }}
        """)
        download_btn.clicked.connect(lambda: self.download_excel(mode))
        layout.addWidget(download_btn)
        
        layout.addStretch()
        
        content.setLayout(layout)
        page.setWidget(content)
        
        # Store references
        if mode == "keluaran":
            self.efaktur_keluaran_filelist = file_list
            self.efaktur_keluaran_status = status
        else:
            self.efaktur_masukan_filelist = file_list
            self.efaktur_masukan_status = status
        
        return page
    
    def build_bupot_page(self):
        """Build BUPOT extraction page"""
        page = QScrollArea()
        page.setWidgetResizable(True)
        page.setStyleSheet(f"background-color: {Config.COLOR_BG_MEDIUM};")
        
        content = QWidget()
        layout = QVBoxLayout()
        layout.setSpacing(20)
        layout.setContentsMargins(40, 20, 40, 20)
        
        # Title
        title = QLabel("📃 BUPOT Extractor")
        title.setFont(QFont("Arial", 24, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet(f"""
            background-color: {Config.COLOR_ACCENT};
            color: black;
            padding: 15px;
            border-radius: 20px;
        """)
        layout.addWidget(title)
        
        # Subtitle
        subtitle = QLabel("Auto-detect BPPU atau BP21 dan ekstrak ke Excel")
        subtitle.setFont(QFont("Arial", 12))
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color: white; padding: 10px;")
        layout.addWidget(subtitle)
        
        # Upload Frame
        upload_frame = QFrame()
        upload_frame.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border: 2px solid {Config.COLOR_ACCENT};
            border-radius: 20px;
            padding: 20px;
        """)
        upload_layout = QVBoxLayout()
        
        upload_label = QLabel("Upload File PDF")
        upload_label.setStyleSheet("color: white; font-size: 14px;")
        upload_label.setAlignment(Qt.AlignCenter)
        upload_layout.addWidget(upload_label)
        
        # File button
        file_button = QPushButton("📁 PILIH FILE PDF")
        file_button.setFixedHeight(45)
        file_button.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_ACCENT};
                color: black;
                border: none;
                border-radius: 22px;
                font-weight: bold;
                font-size: 14px;
            }}
            QPushButton:hover {{
                background-color: #67E8F9;
            }}
        """)
        file_button.clicked.connect(lambda: self.select_files("bupot"))
        upload_layout.addWidget(file_button)
        
        upload_frame.setLayout(upload_layout)
        layout.addWidget(upload_frame)
        
        # File List
        file_frame = QFrame()
        file_frame.setStyleSheet(f"""
            background-color: {Config.COLOR_BG_DARK};
            border-radius: 15px;
            padding: 15px;
        """)
        file_layout = QVBoxLayout()
        
        file_list = QTextEdit()
        file_list.setReadOnly(True)
        file_list.setMinimumHeight(150)
        file_list.setStyleSheet("""
            background-color: #0F172A;
            color: white;
            border: 1px solid #374151;
            border-radius: 8px;
            padding: 10px;
        """)
        file_layout.addWidget(file_list)
        
        file_frame.setLayout(file_layout)
        layout.addWidget(file_frame, 1)
        
        # Process Button
        process_btn = QPushButton("⚡ PROSES SEKARANG")
        process_btn.setFixedHeight(50)
        process_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {Config.COLOR_SUCCESS};
                color: white;
                border: none;
                border-radius: 25px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #059669;
            }}
        """)
        process_btn.clicked.connect(lambda: self.process_extraction("bupot"))
        layout.addWidget(process_btn)
        
        # Status Label
        status = QLabel("Status: Menunggu file...")
        status.setStyleSheet("color: gray; font-size: 12px;")
        status.setAlignment(Qt.AlignCenter)
        layout.addWidget(status)
        
        # Download Button
        download_btn = QPushButton("⬇ DOWNLOAD HASIL EXCEL")
        download_btn.setFixedHeight(50)
        download_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: #A78BFA;
                color: white;
                border: none;
                border-radius: 25px;
                font-weight: bold;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background-color: #C4B5FD;
            }}
        """)
        download_btn.clicked.connect(lambda: self.download_excel("bupot"))
        layout.addWidget(download_btn)
        
        layout.addStretch()
        
        content.setLayout(layout)
        page.setWidget(content)
        
        # Store references
        self.bupot_filelist = file_list
        self.bupot_status = status
        
        return page
    
    def select_files(self, mode):
        """Select PDF files"""
        files, _ = QFileDialog.getOpenFileNames(
            self,
            "Pilih File PDF",
            "",
            "PDF Files (*.pdf)"
        )
        
        if files:
            self.selected_files = files
            
            # Update file list
            if mode == "keluaran":
                file_list = self.efaktur_keluaran_filelist
                status = self.efaktur_keluaran_status
            elif mode == "masukan":
                file_list = self.efaktur_masukan_filelist
                status = self.efaktur_masukan_status
            else:  # bupot
                file_list = self.bupot_filelist
                status = self.bupot_status
            
            file_list.clear()
            for f in files:
                file_list.append(f"✔ {Path(f).name}")
            
            status.setText(f"Status: {len(files)} file siap diproses")
    
    def process_extraction(self, mode):
        """Process PDF extraction"""
        if not hasattr(self, 'selected_files') or not self.selected_files:
            QMessageBox.warning(self, "Peringatan", "Pilih file PDF dulu!")
            return
        
        # Get status label
        if mode == "keluaran":
            status = self.efaktur_keluaran_status
        elif mode == "masukan":
            status = self.efaktur_masukan_status
        else:
            status = self.bupot_status
        
        status.setText("🔄 Memproses dokumen...")
        
        # Run in thread
        thread = threading.Thread(
            target=self._extract_pdfs_thread,
            args=(mode,),
            daemon=True
        )
        thread.start()
    
    def _extract_pdfs_thread(self, mode):
        """Extract PDFs in background thread"""
        dfs = []
        failed = []
        
        for idx, pdf_path in enumerate(self.selected_files, start=1):
            try:
                if mode in ["keluaran", "masukan"]:
                    # Extract using current mapping
                    df = EFakturExtractor.extract(
                        pdf_path,
                        faktur_counter=idx,
                        detect_status=True,
                        field_mapping=self.mapping_controller.get_mapping_snapshot()
                    )
                    
                    # For Masukan, override status
                    if mode == "masukan" and "STATUS" in df.columns:
                        df.loc[df['STATUS'] == 'approved', 'STATUS'] = 'credited'
                else:
                    # BUPOT
                    df = BUPOTExtractor.extract(pdf_path)
                
                dfs.append(df)
                
            except Exception as e:
                failed.append((Path(pdf_path).name, str(e)))
        
        # Combine results
        if dfs:
            self.final_df_global = pd.concat(dfs, ignore_index=True)
            
            success_count = len(self.selected_files) - len(failed)
            
            # Update UI in main thread
            if mode == "keluaran":
                status = self.efaktur_keluaran_status
            elif mode == "masukan":
                status = self.efaktur_masukan_status
            else:
                status = self.bupot_status
            
            QTimer.singleShot(0, lambda: status.setText(
                f"✅ Selesai! {success_count} berhasil, {len(failed)} gagal"
            ))
            
            message = f"✅ {success_count} PDF berhasil diproses."
            if failed:
                message += "\n\n⚠ Beberapa file gagal:\n"
                for fname, error in failed:
                    message += f"- {fname}: {error}\n"
            
            QTimer.singleShot(0, lambda: QMessageBox.information(
                self, "Hasil Ekstraksi", message
            ))
        else:
            if mode == "keluaran":
                status = self.efaktur_keluaran_status
            elif mode == "masukan":
                status = self.efaktur_masukan_status
            else:
                status = self.bupot_status
            
            QTimer.singleShot(0, lambda: status.setText("❌ Tidak ada PDF yang berhasil!"))
            QTimer.singleShot(0, lambda: QMessageBox.critical(
                self, "Error", "Tidak ada PDF yang berhasil diekstrak!"
            ))
    
    def download_excel(self, mode):
        """Download Excel with formatting"""
        if self.final_df_global is None:
            QMessageBox.warning(self, "Peringatan", "Data belum diproses!")
            return
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan Excel",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            
            # Save initial
            self.final_df_global.to_excel(file_path, index=False)
            
            # Reload for formatting
            wb = load_workbook(file_path)
            ws = wb.active
            
            if mode in ["keluaran", "masukan"]:
                # E-FAKTUR FORMATTING
                header_map = {}
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header:
                        header_map[str(header).strip()] = col
                
                # Format cells
                for row_idx in range(2, ws.max_row + 1):
                    for header, col_idx in header_map.items():
                        cell = ws.cell(row=row_idx, column=col_idx)
                        value = cell.value
                        
                        field_info = FieldRegistry.EFAKTUR_FIELDS.get(header)
                        if not field_info:
                            continue
                        
                        if field_info["type"] == "TEXT":
                            cell.value = str(value) if value else ""
                            cell.number_format = "@"
                        elif field_info["type"] == "DATE":
                            dt = EFakturExtractor.parse_tanggal_teks(value)
                            if dt:
                                cell.value = dt
                                cell.number_format = "DD/MM/YYYY"
                        elif field_info["type"] == "NUMBER":
                            cell.value = EFakturExtractor.clean_number(value)
                            cell.number_format = "#,##0"
                
                # Merge cells for HEADER scope fields
                if "NO" in header_map:
                    current_no = None
                    start_row = 2
                    
                    for row_idx in range(2, ws.max_row + 2):
                        if row_idx <= ws.max_row:
                            no_col = header_map.get("NO")
                            current_val = ws.cell(row=row_idx, column=no_col).value if no_col else None
                        else:
                            current_val = None
                        
                        if current_val != current_no and current_no is not None:
                            end_row = row_idx - 1
                            
                            for header, col_idx in header_map.items():
                                field_info = FieldRegistry.EFAKTUR_FIELDS.get(header)
                                if field_info and field_info["merge"] and end_row > start_row:
                                    ws.merge_cells(
                                        start_row=start_row,
                                        start_column=col_idx,
                                        end_row=end_row,
                                        end_column=col_idx
                                    )
                                    ws.cell(row=start_row, column=col_idx).alignment = Alignment(
                                        horizontal='center',
                                        vertical='center',
                                        wrap_text=True
                                    )
                            
                            start_row = row_idx
                        
                        current_no = current_val
            
            else:
                # BUPOT FORMATTING
                for cell in ws['I']:
                    if cell.row > 1:
                        cell.number_format = '@'
                
                for cell in ws['G']:
                    if cell.row > 1:
                        cell.number_format = '#,##0'
                
                for cell in ws['H']:
                    if cell.row > 1:
                        cell.number_format = '#,##0'
                
                for cell in ws['C']:
                    if cell.row > 1:
                        cell.number_format = 'DD/MM/YYYY'
                
                for cell in ws['K']:
                    if cell.row > 1:
                        cell.number_format = 'DD/MM/YYYY'
            
            # Common formatting
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border
            
            wb.save(file_path)
            
            QMessageBox.information(
                self, "Berhasil",
                f"✅ File Excel berhasil disimpan!\n\n{file_path}"
            )
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal menyimpan Excel:\n{e}")
    
    def show_page(self, index):
        """Show page"""
        for i, page in enumerate(self.pages):
            if i == index:
                page.show()
            else:
                page.hide()
        
        # Update button highlights
        buttons = [
            self.btn_home,
            self.btn_mapping,
            self.btn_efaktur_keluaran,
            self.btn_efaktur_masukan,
            self.btn_bupot
        ]
        
        for i, btn in enumerate(buttons):
            if i == index:
                btn.setStyleSheet(f"""
                    QPushButton {{
                        background-color: {Config.COLOR_BUTTON_ACTIVE};
                        border: none;
                        border-radius: 12px;
                        color: white;
                    }}
                    QPushButton:hover {{
                        background-color: {Config.COLOR_BUTTON_HOVER};
                    }}
                """)
            else:
                self.style_sidebar_button(btn)
    
    def set_dark_theme(self):
        """Apply dark theme to application"""
        self.setStyleSheet(f"""
            QMainWindow {{
                background-color: {Config.COLOR_BG_DARK};
            }}
            QLabel {{
                color: white;
            }}
            QScrollArea {{
                border: none;
            }}
            QScrollBar:vertical {{
                background-color: #1E293B;
                width: 12px;
                border-radius: 6px;
            }}
            QScrollBar::handle:vertical {{
                background-color: #475569;
                border-radius: 6px;
            }}
            QScrollBar:horizontal {{
                background-color: #1E293B;
                height: 12px;
                border-radius: 6px;
            }}
            QScrollBar::handle:horizontal {{
                background-color: #475569;
                border-radius: 6px;
            }}
        """)
    
    def copy_to_clipboard(self, text):
        """Copy to clipboard"""
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(self, "Disalin", "Nomor berhasil disalin!")
    
    def _get_config_path(self):
        """Get config file path"""
        if os.name == 'nt':
            base = Path(os.getenv('APPDATA', ''))
        else:
            base = Path.home() / '.config'
        
        config_dir = base / 'SmartPDFExtractor'
        config_dir.mkdir(parents=True, exist_ok=True)
        
        return config_dir / Config.CONFIG_FILENAME
    
    def load_config(self):
        """Load configuration"""
        self._temp_mapping = None
        if self.config_file.exists():
            try:
                with open(self.config_file, "r") as f:
                    data = json.load(f)
                    if "efaktur_mapping" in data:
                        self._temp_mapping = data["efaktur_mapping"]
            except:
                pass
    
    def _apply_loaded_config(self):
        """Apply loaded config to controller"""
        if hasattr(self, '_temp_mapping') and self._temp_mapping:
            self.mapping_controller.load_from_config(self._temp_mapping)
    
    def save_config(self):
        """Save configuration"""
        data = {"efaktur_mapping": self.mapping_controller.field_mapping}
        try:
            with open(self.config_file, "w") as f:
                json.dump(data, f, indent=2)
        except:
            pass

# ============================================================================================
#                                    MAIN
# ============================================================================================
def main():
    app = QApplication(sys.argv)
    
    # Set application-wide font
    app.setFont(QFont("Arial", 10))
    
    window = SmartPDFExtractor()
    window.show()
    
    sys.exit(app.exec())

if __name__ == "__main__":
    main()